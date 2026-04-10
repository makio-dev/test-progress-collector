"""テストデータ生成スクリプト"""
import openpyxl
from datetime import datetime, timedelta
import os

def create_test_file(filepath, sheet_name, test_cases):
    """テスト用Excelファイルを作成"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ヘッダー行（18行目）
    headers = {
        3: "テストID",      # C列
        17: "実施予定",     # Q列
        18: "実施実績",     # R列
        19: "検証予定",     # S列
        20: "検証実績",     # T列
    }
    for col, header in headers.items():
        ws.cell(row=18, column=col, value=header)

    # データ行（19行目から）
    for i, tc in enumerate(test_cases):
        row = 19 + i
        ws.cell(row=row, column=3, value=tc["test_id"])        # C列: テストID
        ws.cell(row=row, column=17, value=tc.get("jisshi_yotei"))   # Q列
        ws.cell(row=row, column=18, value=tc.get("jisshi_jisseki")) # R列
        ws.cell(row=row, column=19, value=tc.get("kensho_yotei"))   # S列
        ws.cell(row=row, column=20, value=tc.get("kensho_jisseki")) # T列

    wb.save(filepath)
    print(f"Created: {filepath}")


def main():
    input_dir = "/Users/takuminishimaki/Develop/makio-dev/test-progress-collector/input"
    os.makedirs(input_dir, exist_ok=True)

    # サブフォルダも作成
    os.makedirs(os.path.join(input_dir, "subdir"), exist_ok=True)

    # 基準日（今日）
    today = datetime.now()

    # 日付ヘルパー
    def d(offset):
        """今日からのオフセット日付を返す"""
        return today + timedelta(days=offset)

    # === ファイル1: オンラインチーム (-O-) ===
    create_test_file(
        os.path.join(input_dir, "TEST-O-001_オンライン機能.xlsx"),
        "ITB_ログイン",
        [
            {"test_id": "TC001", "jisshi_yotei": d(-5), "jisshi_jisseki": d(-5), "kensho_yotei": d(-3), "kensho_jisseki": d(-3)},
            {"test_id": "TC002", "jisshi_yotei": d(-5), "jisshi_jisseki": d(-4), "kensho_yotei": d(-3), "kensho_jisseki": d(-2)},
            {"test_id": "TC003", "jisshi_yotei": d(-4), "jisshi_jisseki": d(-4), "kensho_yotei": d(-2), "kensho_jisseki": d(-1)},
            {"test_id": "TC004", "jisshi_yotei": d(-3), "jisshi_jisseki": d(-3), "kensho_yotei": d(-1), "kensho_jisseki": d(0)},
            {"test_id": "TC005", "jisshi_yotei": d(-2), "jisshi_jisseki": d(-2), "kensho_yotei": d(0), "kensho_jisseki": None},  # 検証未完了
            {"test_id": "TC006", "jisshi_yotei": d(-1), "jisshi_jisseki": d(0), "kensho_yotei": d(1), "kensho_jisseki": None},   # 実施遅延
            {"test_id": "TC007", "jisshi_yotei": d(0), "jisshi_jisseki": None, "kensho_yotei": d(2), "kensho_jisseki": None},    # 実施未完了
            {"test_id": "TC008", "jisshi_yotei": d(1), "jisshi_jisseki": None, "kensho_yotei": d(3), "kensho_jisseki": None},    # 将来予定
            {"test_id": "TC009", "jisshi_yotei": d(2), "jisshi_jisseki": None, "kensho_yotei": d(4), "kensho_jisseki": None},    # 将来予定
            {"test_id": "TC010", "jisshi_yotei": d(3), "jisshi_jisseki": None, "kensho_yotei": d(5), "kensho_jisseki": None},    # 将来予定
        ]
    )

    # === ファイル2: バッチチーム (-B-) ===
    create_test_file(
        os.path.join(input_dir, "TEST-B-001_バッチ処理.xlsx"),
        "ITB_日次処理",
        [
            {"test_id": "BT001", "jisshi_yotei": d(-4), "jisshi_jisseki": d(-4), "kensho_yotei": d(-2), "kensho_jisseki": d(-2)},
            {"test_id": "BT002", "jisshi_yotei": d(-3), "jisshi_jisseki": d(-3), "kensho_yotei": d(-1), "kensho_jisseki": d(-1)},
            {"test_id": "BT003", "jisshi_yotei": d(-2), "jisshi_jisseki": d(-1), "kensho_yotei": d(0), "kensho_jisseki": None},  # 実施遅延
            {"test_id": "BT004", "jisshi_yotei": d(-1), "jisshi_jisseki": None, "kensho_yotei": d(1), "kensho_jisseki": None},   # 実施未完了
            {"test_id": "BT005", "jisshi_yotei": d(0), "jisshi_jisseki": None, "kensho_yotei": d(2), "kensho_jisseki": None},    # 今日予定
        ]
    )

    # === ファイル3: 基盤チーム (-I-) サブフォルダ内 ===
    create_test_file(
        os.path.join(input_dir, "subdir", "TEST-I-001_基盤共通.xlsx"),
        "ITB-DB接続",
        [
            {"test_id": "IT001", "jisshi_yotei": d(-3), "jisshi_jisseki": d(-3), "kensho_yotei": d(-1), "kensho_jisseki": d(-1)},
            {"test_id": "IT002", "jisshi_yotei": d(-2), "jisshi_jisseki": d(-2), "kensho_yotei": d(0), "kensho_jisseki": d(0)},
            {"test_id": "IT003", "jisshi_yotei": d(-1), "jisshi_jisseki": d(-1), "kensho_yotei": d(1), "kensho_jisseki": None},
        ]
    )

    # === ファイル4: 運用チーム (-U-) サブフォルダ内 ===
    create_test_file(
        os.path.join(input_dir, "subdir", "TEST-U-001_運用監視.xlsx"),
        "ITB_監視機能",
        [
            {"test_id": "UT001", "jisshi_yotei": d(-2), "jisshi_jisseki": d(-2), "kensho_yotei": d(0), "kensho_jisseki": d(0)},
            {"test_id": "UT002", "jisshi_yotei": d(-1), "jisshi_jisseki": d(-1), "kensho_yotei": d(1), "kensho_jisseki": None},
            {"test_id": "UT003", "jisshi_yotei": d(0), "jisshi_jisseki": None, "kensho_yotei": d(2), "kensho_jisseki": None},
            {"test_id": "UT004", "jisshi_yotei": d(1), "jisshi_jisseki": None, "kensho_yotei": d(3), "kensho_jisseki": None},
        ]
    )

    # === ファイル5: その他（パターンなし） ===
    create_test_file(
        os.path.join(input_dir, "TEST_その他機能.xlsx"),
        "ITB_その他",
        [
            {"test_id": "OT001", "jisshi_yotei": d(-1), "jisshi_jisseki": d(-1), "kensho_yotei": d(1), "kensho_jisseki": None},
            {"test_id": "OT002", "jisshi_yotei": d(0), "jisshi_jisseki": None, "kensho_yotei": d(2), "kensho_jisseki": None},
        ]
    )

    print("\n=== 単一フォルダ テストデータ作成完了 ===")
    print(f"入力フォルダ: {input_dir}")
    print("作成ファイル:")
    print("  - TEST-O-001_オンライン機能.xlsx (オンラインチーム, 10件)")
    print("  - TEST-B-001_バッチ処理.xlsx (バッチチーム, 5件)")
    print("  - subdir/TEST-I-001_基盤共通.xlsx (基盤チーム, 3件)")
    print("  - subdir/TEST-U-001_運用監視.xlsx (運用チーム, 4件)")
    print("  - TEST_その他機能.xlsx (その他, 2件)")
    print("\n合計: 24件のテストケース")

    # ===================================================================
    #  複数フォルダテスト用データ
    # ===================================================================
    base_dir = os.path.dirname(input_dir)
    teamA_dir = os.path.join(base_dir, "input_multi", "teamA")
    teamB_dir = os.path.join(base_dir, "input_multi", "teamB")
    os.makedirs(teamA_dir, exist_ok=True)
    os.makedirs(teamB_dir, exist_ok=True)

    # === teamA: オンライン + バッチ ===
    create_test_file(
        os.path.join(teamA_dir, "ITB-O-101_ログイン認証.xlsx"),
        "ITB-ログイン認証",
        [
            {"test_id": "MA-O-001", "jisshi_yotei": d(-6), "jisshi_jisseki": d(-6), "kensho_yotei": d(-4), "kensho_jisseki": d(-4)},
            {"test_id": "MA-O-002", "jisshi_yotei": d(-5), "jisshi_jisseki": d(-5), "kensho_yotei": d(-3), "kensho_jisseki": d(-3)},
            {"test_id": "MA-O-003", "jisshi_yotei": d(-4), "jisshi_jisseki": d(-4), "kensho_yotei": d(-2), "kensho_jisseki": d(-1)},
            {"test_id": "MA-O-004", "jisshi_yotei": d(-3), "jisshi_jisseki": d(-2), "kensho_yotei": d(0), "kensho_jisseki": None},
            {"test_id": "MA-O-005", "jisshi_yotei": d(-1), "jisshi_jisseki": None, "kensho_yotei": d(1), "kensho_jisseki": None},
        ]
    )

    create_test_file(
        os.path.join(teamA_dir, "ITB-B-101_日次バッチ.xlsx"),
        "ITB-日次バッチ",
        [
            {"test_id": "MA-B-001", "jisshi_yotei": d(-4), "jisshi_jisseki": d(-4), "kensho_yotei": d(-2), "kensho_jisseki": d(-2)},
            {"test_id": "MA-B-002", "jisshi_yotei": d(-3), "jisshi_jisseki": d(-3), "kensho_yotei": d(-1), "kensho_jisseki": d(-1)},
            {"test_id": "MA-B-003", "jisshi_yotei": d(-1), "jisshi_jisseki": None, "kensho_yotei": d(1), "kensho_jisseki": None},
        ]
    )

    # === teamB: 基盤 + 運用 ===
    create_test_file(
        os.path.join(teamB_dir, "ITB-I-101_DB接続テスト.xlsx"),
        "ITB-DB接続",
        [
            {"test_id": "MB-I-001", "jisshi_yotei": d(-5), "jisshi_jisseki": d(-5), "kensho_yotei": d(-3), "kensho_jisseki": d(-3)},
            {"test_id": "MB-I-002", "jisshi_yotei": d(-3), "jisshi_jisseki": d(-3), "kensho_yotei": d(-1), "kensho_jisseki": d(-1)},
            {"test_id": "MB-I-003", "jisshi_yotei": d(-1), "jisshi_jisseki": d(0), "kensho_yotei": d(1), "kensho_jisseki": None},
            {"test_id": "MB-I-004", "jisshi_yotei": d(1), "jisshi_jisseki": None, "kensho_yotei": d(3), "kensho_jisseki": None},
        ]
    )

    create_test_file(
        os.path.join(teamB_dir, "ITB-U-101_監視アラート.xlsx"),
        "ITB-監視アラート",
        [
            {"test_id": "MB-U-001", "jisshi_yotei": d(-4), "jisshi_jisseki": d(-4), "kensho_yotei": d(-2), "kensho_jisseki": d(-2)},
            {"test_id": "MB-U-002", "jisshi_yotei": d(-2), "jisshi_jisseki": d(-2), "kensho_yotei": d(0), "kensho_jisseki": None},
            {"test_id": "MB-U-003", "jisshi_yotei": d(0), "jisshi_jisseki": None, "kensho_yotei": d(2), "kensho_jisseki": None},
            {"test_id": "MB-U-004", "jisshi_yotei": d(2), "jisshi_jisseki": None, "kensho_yotei": d(4), "kensho_jisseki": None},
        ]
    )

    print("\n=== 複数フォルダ テストデータ作成完了 ===")
    print(f"teamA: {teamA_dir}")
    print("  - ITB-O-101_ログイン認証.xlsx (オンライン, 5件)")
    print("  - ITB-B-101_日次バッチ.xlsx (バッチ, 3件)")
    print(f"teamB: {teamB_dir}")
    print("  - ITB-I-101_DB接続テスト.xlsx (基盤, 4件)")
    print("  - ITB-U-101_監視アラート.xlsx (運用, 4件)")
    print("\n複数フォルダ合計: 16件のテストケース")
    print("\n--- 複数フォルダでの実行例 ---")
    print(f"python aggregate_test_results.py {teamA_dir} {teamB_dir} -o ./output/test_multi.xlsx")


if __name__ == "__main__":
    main()
