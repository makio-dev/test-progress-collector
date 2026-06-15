"""PB曲線（信頼度成長曲線）生成スクリプト

test-progress-collector が扱うテストケース／欠陥データから、
PB曲線（P系=テスト消化バーンダウン ＋ B系=欠陥検出）を1枚に重ねた
Excelを自動生成する独立ツール。

データ抽出は aggregate_test_results.py の収集関数を再利用する。
生成するExcelは「数式駆動」で、入力面シート（パラメータ／入力データ／欠陥データ）を
Excel上で編集すると、計算シリーズとグラフが自動的に再計算される。

使い方:
    python generate_pb_curve.py <folder...> -o ./output/pb_curve.xlsx \
        [--pivot-date YYYY-MM-DD] [--start-date ...] [--end-date ...] \
        [--b-final-rate 0.0105] [--b-lower-rate 0.0035] [--b-upper-rate 0.0213] \
        [--forecast-mult 0.0224] [--total-case N] \
        [--defect-online path] [--defect-batch path] \
        [--defect-infra path] [--defect-ops path] [--no-subfolders]

EXE化（既存ツールと同名・別build）:
    pyinstaller --onefile --windowed --name aggregate_test_results \
        --distpath dist_pb generate_pb_curve.py
"""
import argparse
import os
import sys
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, AreaChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont,
)

# 既存システムの収集ロジック・ユーティリティを再利用（重複実装しない）
import aggregate_test_results as agg

# ===================================================================
#  デフォルト係数（写真のパラメータ値）
# ===================================================================
DEFAULT_B_FINAL_RATE = 0.0105   # B系最終計画 = テストケース数 × この値
DEFAULT_B_LOWER_RATE = 0.0035   # B系目標帯（下限）
DEFAULT_B_UPPER_RATE = 0.0213   # B系目標帯（上限）

# パラメータシートの固定セル参照（数式で他シートから参照する）
# B系の目標値は「テストケース数×係数」を1セルに統合（係数と値の二重持ちはしない）
P_START = "パラメータ!$B$3"
P_END = "パラメータ!$B$4"
P_PIVOT = "パラメータ!$B$5"
P_TOTAL = "パラメータ!$B$6"
P_PLAN_FINAL = "パラメータ!$B$7"   # =テストケース数×最終計画係数
P_BAND_LOWER = "パラメータ!$B$8"   # =テストケース数×下限係数
P_BAND_UPPER = "パラメータ!$B$9"   # =テストケース数×上限係数
P_FORECAST = "パラメータ!$B$10"    # 予測倍率

# スタイル
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="游ゴシック", size=14, bold=True)
KEY_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN = agg.THIN_BORDER

# グラフ系列の色（互いに明確に区別できる配色）
COLOR_P_PLAN = "E8000B"      # P未実施（計画）赤・破線
COLOR_P_ACTUAL = "000000"    # P未実施（実績）黒・実線（太）
COLOR_B_ACTUAL = "7030A0"    # B実績 紫・実線
COLOR_B_PLAN = "00B050"      # B計画 緑・破線
COLOR_B_FORECAST = "FF8C00"  # B予測 オレンジ・点線
COLOR_BAND = "FCE4EC"        # B目標レンジ 薄ピンク


# ===================================================================
#  データ準備
# ===================================================================
def _parse_cli_date(s):
    """YYYY-MM-DD / YYYY/MM/DD / YYYYMMDD を date に変換"""
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"日付形式が不正です: {s}")


def _rec_date(val):
    """レコードの日付値（'YYYY/MM/DD' 文字列 or date or None）を date に正規化"""
    return agg._to_date_obj(val)


def build_cases(records):
    """テストケースの入力リストを構築（P系の元データ）

    Returns: list of dict(テストID, チーム名, 予定:date|None, 実績:date|None)
    """
    cases = []
    for r in records:
        cases.append({
            "テストID": r.get("テストID", ""),
            "チーム名": r.get("チーム名", ""),
            "予定": _rec_date(r.get("実施者_予定")),
            "実績": _rec_date(r.get("実施者_実績")),
        })
    return cases


def build_defects(defect_detail_records):
    """欠陥の入力リストを構築（B系の元データ）

    Returns: list of dict(欠陥ID, チーム名, 発見日:date|None, 件名)
    """
    defects = []
    for r in defect_detail_records:
        defects.append({
            "欠陥ID": r.get("欠陥ID", ""),
            "チーム名": r.get("チーム名", ""),
            "発見日": _rec_date(r.get("発見日")),
            "件名": r.get("件名", ""),
        })
    return defects


def resolve_period(cases, pivot, start_override, end_override):
    """開始日・終了日を決定（予定/実績の全日付から min/max、基準日を内包）"""
    all_dates = []
    for c in cases:
        if c["予定"]:
            all_dates.append(c["予定"])
        if c["実績"]:
            all_dates.append(c["実績"])
    if not all_dates:
        raise ValueError("テストケースに有効な実施予定日・実績日が1件もありません。")

    start = start_override or min(all_dates)
    end = end_override or max(all_dates)
    # 基準日を期間内に収める
    if pivot < start:
        pivot = start
    if pivot > end:
        end = pivot
    if end < start:
        end = start
    return start, end, pivot


# ===================================================================
#  シート書き込み
# ===================================================================
def _style_header_row(ws, row, ncol):
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN


def write_param_sheet(ws, start, end, pivot, total_case,
                      b_final_rate, b_lower_rate, b_upper_rate,
                      forecast_mult, pivot_row):
    """パラメータシート（Key/Value/日本語/解説）"""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "パラメータ"
    ws["A1"].font = TITLE_FONT

    headers = ["Key", "Value", "日本語", "解説"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, 4)

    # Forecast_Mult: 指定があれば固定値、なければ実績から自動算出する数式
    if forecast_mult is None:
        forecast_value = (f"=IF(P_シリーズ!$E${pivot_row}=0,0,"
                          f"B_シリーズ!$C${pivot_row}/P_シリーズ!$E${pivot_row})")
    else:
        forecast_value = forecast_mult

    # B系の目標値は「テストケース数(B6)×係数」を1セルに統合（係数は数式内に保持）。
    # 係数を変えたい場合はこのセルの数式の数値部分を直接編集する。
    rows = [
        ("StartDate", start, "開始日", "分析対象期間の開始日"),
        ("EndDate", end, "終了日", "分析対象期間の終了日"),
        ("PivotDate", pivot, "基準日（ピボット日）", "PB曲線の評価基準日。実績はこの日まで描画する"),
        ("TotalCase", total_case, "テストケース総数", "P系（消化バーンダウン）の対象件数"),
        ("B_Plan_Final", f"={P_TOTAL}*{b_final_rate}", "B系最終計画値",
         f"テストケース数×{b_final_rate}（最終到達の計画欠陥数）"),
        ("B_Band_Lower", f"={P_TOTAL}*{b_lower_rate}", "B系目標帯（下限）",
         f"テストケース数×{b_lower_rate}（目標レンジ下限）"),
        ("B_Band_Upper", f"={P_TOTAL}*{b_upper_rate}", "B系目標帯（上限）",
         f"テストケース数×{b_upper_rate}（目標レンジ上限）"),
        ("Forecast_Mult", forecast_value, "予測倍率（Forecast係数）",
         "基準日以降の欠陥発生見込み（欠陥/ケース）"),
    ]

    for i, (key, value, jp, desc) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=key).font = Font(name="游ゴシック", size=10, bold=True)
        ws.cell(row=r, column=1).fill = KEY_FILL
        vc = ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=jp)
        ws.cell(row=r, column=4, value=desc)
        # 書式
        if isinstance(value, date) and not isinstance(value, datetime):
            vc.number_format = "yyyy/mm/dd"
        elif key == "Forecast_Mult":
            vc.number_format = "0.0000"
        elif key in ("B_Plan_Final", "B_Band_Lower", "B_Band_Upper"):
            vc.number_format = "0.00"
        elif key == "TotalCase":
            vc.number_format = "#,##0"
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = THIN
            ws.cell(row=r, column=col).alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 50


def write_input_sheet(ws, cases):
    """入力データシート（P系の元データ＝編集可能）。戻り値: (data_start, data_end)"""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "入力データ（テストケース実施）"
    ws["A1"].font = TITLE_FONT
    headers = ["No", "テストID", "チーム名", "実施者_予定", "実施者_実績"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))

    data_start = 3
    for i, c in enumerate(cases):
        r = data_start + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=c["テストID"])
        ws.cell(row=r, column=3, value=c["チーム名"])
        if c["予定"]:
            d = ws.cell(row=r, column=4, value=c["予定"])
            d.number_format = "yyyy/mm/dd"
        if c["実績"]:
            d = ws.cell(row=r, column=5, value=c["実績"])
            d.number_format = "yyyy/mm/dd"
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = THIN
    data_end = data_start + len(cases) - 1 if cases else data_start
    for col, w in zip("ABCDE", [6, 18, 12, 14, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    return data_start, max(data_end, data_start)


def write_defect_sheet(ws, defects):
    """欠陥データシート（B系の元データ＝編集可能・行削除で除外可）。戻り値: (data_start, data_end)"""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "欠陥データ（発見日）"
    ws["A1"].font = TITLE_FONT
    ws["D1"] = "※ 算出から除外したい欠陥は行ごと削除してください（グラフが自動再計算されます）"
    ws["D1"].font = Font(name="游ゴシック", size=9, color="C00000")
    headers = ["No", "欠陥ID", "チーム名", "発見日", "件名"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))

    data_start = 3
    for i, d in enumerate(defects):
        r = data_start + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=d["欠陥ID"])
        ws.cell(row=r, column=3, value=d["チーム名"])
        if d["発見日"]:
            dc = ws.cell(row=r, column=4, value=d["発見日"])
            dc.number_format = "yyyy/mm/dd"
        ws.cell(row=r, column=5, value=d["件名"])
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = THIN
    data_end = data_start + len(defects) - 1 if defects else data_start
    for col, w in zip("ABCDE", [6, 14, 12, 14, 40]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    return data_start, max(data_end, data_start)


def write_p_series_sheet(ws, dates, input_start, input_end):
    """P_シリーズ（日次の計画/実績バーンダウン）。すべて数式駆動。"""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "P_シリーズ（テスト消化）"
    ws["A1"].font = TITLE_FONT
    headers = ["日付", "P予定日次", "P実績日次", "P予定累計", "P実績累計",
               "P未実施計画", "P未実施実績"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))

    plan_col = f"入力データ!$D${input_start}:$D${input_end}"
    act_col = f"入力データ!$E${input_start}:$E${input_end}"

    data_start = 3
    for i, dt in enumerate(dates):
        r = data_start + i
        dc = ws.cell(row=r, column=1, value=dt)
        dc.number_format = "mm/dd"
        # B: P予定日次, C: P実績日次
        ws.cell(row=r, column=2, value=f"=COUNTIFS({plan_col},$A{r})")
        ws.cell(row=r, column=3, value=f"=COUNTIFS({act_col},$A{r})")
        # D: P予定累計, E: P実績累計
        ws.cell(row=r, column=4, value=f"=SUM($B${data_start}:$B{r})")
        ws.cell(row=r, column=5, value=f"=SUM($C${data_start}:$C{r})")
        # F: P未実施計画 = TotalCase - 予定累計
        ws.cell(row=r, column=6, value=f"={P_TOTAL}-$D{r}")
        # G: P未実施実績 = 基準日まで TotalCase - 実績累計、以降 NA()
        ws.cell(row=r, column=7, value=f"=IF($A{r}<={P_PIVOT},{P_TOTAL}-$E{r},NA())")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN
            if col >= 2:
                cell.number_format = "#,##0"
    data_end = data_start + len(dates) - 1
    ws.column_dimensions["A"].width = 9
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 11
    ws.freeze_panes = "B3"
    return data_start, data_end


def write_b_series_sheet(ws, dates, defect_start, defect_end, pivot_row):
    """B_シリーズ（日次の実績/計画/目標帯/予測）。すべて数式駆動。

    P_シリーズと同じ行配置（行3起点・同じ日付）であることが前提。
    """
    ws.sheet_view.showGridLines = False
    ws["A1"] = "B_シリーズ（欠陥検出）"
    ws["A1"].font = TITLE_FONT
    headers = ["日付", "B検出日次", "B実績累計", "進捗率", "B計画",
               "B帯下限", "B帯幅", "B予測"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header_row(ws, 2, len(headers))

    det_col = f"欠陥データ!$D${defect_start}:$D${defect_end}"

    data_start = 3
    for i, dt in enumerate(dates):
        r = data_start + i
        dc = ws.cell(row=r, column=1, value=dt)
        dc.number_format = "mm/dd"
        # B: B検出日次
        ws.cell(row=r, column=2, value=f"=COUNTIFS({det_col},$A{r})")
        # C: B実績累計（基準日まで、以降 NA()）
        ws.cell(row=r, column=3,
                value=f'=IF($A{r}<={P_PIVOT},COUNTIFS({det_col},"<="&$A{r}),NA())')
        # D: 進捗率 = P予定累計 / TotalCase
        ws.cell(row=r, column=4,
                value=f"=IF({P_TOTAL}=0,0,P_シリーズ!$D{r}/{P_TOTAL})")
        # E: B計画 = B_Plan_Final × 進捗率
        ws.cell(row=r, column=5, value=f"={P_PLAN_FINAL}*$D{r}")
        # F: B帯下限 = B_Band_Lower × 進捗率
        ws.cell(row=r, column=6, value=f"={P_BAND_LOWER}*$D{r}")
        # G: B帯幅 = (B_Band_Upper - B_Band_Lower) × 進捗率（ピンク帯の積み上げ用）
        ws.cell(row=r, column=7, value=f"=({P_BAND_UPPER}-{P_BAND_LOWER})*$D{r}")
        # H: B予測（基準日以降のみ）= 基準日の実績累計 + 予測倍率 ×（予定累計 - 基準日の予定累計）
        ws.cell(row=r, column=8,
                value=(f"=IF($A{r}<={P_PIVOT},NA(),"
                       f"$C${pivot_row}+{P_FORECAST}*(P_シリーズ!$D{r}-P_シリーズ!$D${pivot_row}))"))
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN
            if col == 4:
                cell.number_format = "0.0%"
            elif col >= 2:
                cell.number_format = "#,##0.0"
    data_end = data_start + len(dates) - 1
    ws.column_dimensions["A"].width = 9
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 11
    ws.freeze_panes = "B3"
    return data_start, data_end


def _line_series(chart, ws, col, start, end, name, color,
                 dash=None, width=19050):
    """1系列をチャートに追加して書式設定（width: EMU, 9525=0.75pt）"""
    ref = Reference(ws, min_col=col, min_row=start, max_row=end)
    chart.add_data(ref, titles_from_data=False)
    s = chart.series[-1]
    s.tx = SeriesLabel(v=name)
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width = width
    if dash:
        s.graphicalProperties.line.dashStyle = dash
    s.marker.symbol = "none"
    s.smooth = False
    return s


def _axis_text(sz):
    """軸・凡例・タイトル用の RichText（フォント）を生成"""
    font = DrawingFont(typeface="ＭＳ Ｐゴシック")
    cp = CharacterProperties(latin=font, sz=sz)
    return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])


def write_graph_sheet(ws, p_ws, b_ws, p_start, p_end, b_start, b_end):
    """グラフシート（二軸：P系=左、B系=右、ピンク帯）

    描画順（z順）= 結合順。帯を最背面、その上にB系ライン、最前面にP系ラインを重ねる。
    ベースチャート=帯（第2軸=右）、P系ライン=主軸（左）。
    """
    ws.sheet_view.showGridLines = False
    ws["A1"] = "PB曲線"
    ws["A1"].font = TITLE_FONT

    n_days = p_end - p_start + 1
    # X軸ラベルの間引き（約14本に収まるよう日数で間引く）
    skip = max(1, round(n_days / 14))

    cats = Reference(p_ws, min_col=1, min_row=p_start, max_row=p_end)

    # --- ピンク帯（積み上げエリア：下限=透明 + 帯幅=ピンク）＝ベース・最背面・第2軸（右） ---
    band = AreaChart()
    band.title = "PB曲線（テスト消化 × 欠陥検出）"
    band.title.txPr = _axis_text(1100)
    band.style = 10
    band.height = 13
    band.width = 30
    band.roundedCorners = False
    band.grouping = "stacked"
    band.overlap = 100
    lower_ref = Reference(b_ws, min_col=6, min_row=b_start, max_row=b_end)  # F: B帯下限
    width_ref = Reference(b_ws, min_col=7, min_row=b_start, max_row=b_end)  # G: B帯幅
    band.add_data(lower_ref, titles_from_data=False)
    band.add_data(width_ref, titles_from_data=False)
    band.set_categories(cats)
    # 下限は塗りなし
    band.series[0].graphicalProperties.noFill = True
    band.series[0].graphicalProperties.line.noFill = True
    band.series[0].tx = SeriesLabel(v="（下限）")
    # 帯はピンク塗り
    band.series[1].graphicalProperties.solidFill = COLOR_BAND
    band.series[1].graphicalProperties.line.noFill = True
    band.series[1].tx = SeriesLabel(v="B目標レンジ")

    # X軸（共有・カテゴリ＝日付）: ラベル間引き・短い日付書式・薄いフォント
    band.x_axis.delete = False
    band.x_axis.number_format = "m/d"
    band.x_axis.tickLblPos = "low"
    band.x_axis.tickLblSkip = skip
    band.x_axis.tickMarkSkip = skip
    band.x_axis.txPr = _axis_text(800)
    band.x_axis.majorGridlines = None
    # 右Y軸（B系）
    band.y_axis.axId = 200
    band.y_axis.title = "欠陥数（B系）"
    band.y_axis.txPr = _axis_text(800)
    band.y_axis.crosses = "max"
    band.y_axis.delete = False
    band.y_axis.majorGridlines = None

    # 凡例（下・小さめフォント）。透明な「（下限）」系列は凡例から除外
    band.legend.position = "b"
    band.legend.overlay = False
    band.legend.txPr = _axis_text(800)
    band.legend.legendEntry = [LegendEntry(idx=0, delete=True)]

    # --- B系ライン（第2軸・右） ---
    b_chart = LineChart()
    _line_series(b_chart, b_ws, 3, b_start, b_end, "B実績(〜基準日)", COLOR_B_ACTUAL, width=28575)  # C 実績=太
    _line_series(b_chart, b_ws, 5, b_start, b_end, "B計画", COLOR_B_PLAN, dash="sysDash", width=15875)  # E
    _line_series(b_chart, b_ws, 8, b_start, b_end, "B予測", COLOR_B_FORECAST, dash="sysDot", width=19050)  # H
    b_chart.set_categories(cats)
    b_chart.y_axis.axId = 200
    b_chart.y_axis.crosses = "max"

    # --- P系ライン（主軸・左・最前面） ---
    p_chart = LineChart()
    _line_series(p_chart, p_ws, 6, p_start, p_end, "P未実施(計画)", COLOR_P_PLAN, dash="sysDash", width=19050)  # F
    _line_series(p_chart, p_ws, 7, p_start, p_end, "P未実施(実績〜基準日)", COLOR_P_ACTUAL, width=28575)        # G 実績=太
    p_chart.set_categories(cats)
    p_chart.y_axis.axId = 100
    p_chart.y_axis.title = "未実施テストケース数（P系）"
    p_chart.y_axis.txPr = _axis_text(800)
    p_chart.y_axis.delete = False
    # 左Y軸の目盛り線のみ薄いグレーで表示
    p_chart.y_axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9", w=9525))
    )

    # プロット領域を明示（凡例・軸ラベルのための余白確保）
    band.layout = Layout(manualLayout=ManualLayout(
        layoutTarget="inner", xMode="edge", yMode="edge", wMode="factor", hMode="factor",
        x=0.07, y=0.12, w=0.86, h=0.70,
    ))

    # 帯(背面) → B系ライン → P系ライン(前面) の順に重ねる
    band += b_chart
    band += p_chart

    ws.add_chart(band, "A3")


# ===================================================================
#  メイン処理
# ===================================================================
def generate(folder_paths, output_path, include_subfolders=True, defect_files=None,
             pivot_date=None, start_date=None, end_date=None,
             b_final_rate=DEFAULT_B_FINAL_RATE, b_lower_rate=DEFAULT_B_LOWER_RATE,
             b_upper_rate=DEFAULT_B_UPPER_RATE, forecast_mult=None, total_case=None):
    """PB曲線Excelを生成する"""
    print("=== PB曲線生成 ===")
    print("テストケースを収集中...")
    records = agg.collect_data(folder_paths, cache_file=None,
                               include_subfolders=include_subfolders)
    cases = build_cases(records)
    print(f"  テストケース: {len(cases)}件")

    defect_detail_records = []
    if defect_files:
        print("欠陥詳細を収集中...")
        defect_detail_records = agg.collect_defect_detail_data(defect_files)
    defects = build_defects(defect_detail_records)
    print(f"  欠陥: {len(defects)}件")

    # 基準日（前営業日）
    if pivot_date is None:
        pivot_date = agg.get_previous_business_day(datetime.now(), agg.DEFAULT_HOLIDAYS).date()

    start, end, pivot = resolve_period(cases, pivot_date, start_date, end_date)
    dates = agg.generate_date_range(start, end)
    if total_case is None:
        total_case = len(cases)

    # 基準日の行番号（P/B シリーズ共通、行3起点）
    pivot_idx = (pivot - start).days
    pivot_idx = max(0, min(pivot_idx, len(dates) - 1))
    pivot_row = 3 + pivot_idx

    print(f"  期間: {start} 〜 {end}（{len(dates)}日） / 基準日: {pivot}")
    print(f"  TotalCase={total_case}  係数: 最終={b_final_rate} 下限={b_lower_rate} 上限={b_upper_rate}")

    wb = openpyxl.Workbook()
    ws_param = wb.active
    ws_param.title = "パラメータ"
    ws_input = wb.create_sheet("入力データ")
    ws_defect = wb.create_sheet("欠陥データ")
    ws_p = wb.create_sheet("P_シリーズ")
    ws_b = wb.create_sheet("B_シリーズ")
    ws_graph = wb.create_sheet("グラフ")

    write_param_sheet(ws_param, start, end, pivot, total_case,
                      b_final_rate, b_lower_rate, b_upper_rate, forecast_mult, pivot_row)
    in_start, in_end = write_input_sheet(ws_input, cases)
    df_start, df_end = write_defect_sheet(ws_defect, defects)
    p_start, p_end = write_p_series_sheet(ws_p, dates, in_start, in_end)
    b_start, b_end = write_b_series_sheet(ws_b, dates, df_start, df_end, pivot_row)
    write_graph_sheet(ws_graph, ws_p, ws_b, p_start, p_end, b_start, b_end)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\n出力完了: {output_path}")
    return output_path


def run_gui():
    """引数なし起動（EXEダブルクリック等）向けの簡易GUI。

    フォルダ・出力先・欠陥ファイル・基準日・B系係数を指定して生成する。
    """
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.title("PB曲線 生成")
    root.geometry("680x520")

    pad = {"padx": 6, "pady": 3}
    vars_ = {}

    def row(label, r, default="", browse=None):
        tk.Label(root, text=label, anchor="w", width=18).grid(row=r, column=0, sticky="w", **pad)
        v = tk.StringVar(value=default)
        e = tk.Entry(root, textvariable=v, width=58)
        e.grid(row=r, column=1, sticky="w", **pad)
        if browse:
            tk.Button(root, text="参照", command=lambda: browse(v)).grid(row=r, column=2, **pad)
        return v

    def pick_dir(v):
        d = filedialog.askdirectory()
        if d:
            v.set(d)

    def pick_save(v):
        f = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                         filetypes=[("Excel", "*.xlsx")])
        if f:
            v.set(f)

    def pick_file(v):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm"), ("すべて", "*.*")])
        if f:
            v.set(f)

    r = 0
    tk.Label(root, text="PB曲線（信頼度成長曲線）生成", font=("", 13, "bold")).grid(
        row=r, column=0, columnspan=3, sticky="w", padx=6, pady=8); r += 1
    vars_["folder"] = row("入力フォルダ *", r, browse=pick_dir); r += 1
    vars_["output"] = row("出力ファイル", r, default="pb_curve.xlsx", browse=pick_save); r += 1
    vars_["d_online"] = row("欠陥一覧(オンライン)", r, browse=pick_file); r += 1
    vars_["d_batch"] = row("欠陥一覧(バッチ)", r, browse=pick_file); r += 1
    vars_["d_infra"] = row("欠陥一覧(基盤)", r, browse=pick_file); r += 1
    vars_["d_ops"] = row("欠陥一覧(運用)", r, browse=pick_file); r += 1
    vars_["pivot"] = row("基準日(空欄=前営業日)", r, default=""); r += 1
    vars_["start"] = row("開始日(空欄=自動)", r, default=""); r += 1
    vars_["end"] = row("終了日(空欄=自動)", r, default=""); r += 1
    vars_["fr"] = row("B系 最終計画係数", r, default=str(DEFAULT_B_FINAL_RATE)); r += 1
    vars_["lr"] = row("B系 下限係数", r, default=str(DEFAULT_B_LOWER_RATE)); r += 1
    vars_["ur"] = row("B系 上限係数", r, default=str(DEFAULT_B_UPPER_RATE)); r += 1
    sub = tk.BooleanVar(value=True)
    tk.Checkbutton(root, text="サブフォルダも探索する", variable=sub).grid(
        row=r, column=1, sticky="w", **pad); r += 1

    def on_run():
        folder = vars_["folder"].get().strip()
        if not folder:
            messagebox.showerror("エラー", "入力フォルダを指定してください。")
            return
        defect_files = {}
        for key, team in (("d_online", "オンライン"), ("d_batch", "バッチ"),
                          ("d_infra", "基盤"), ("d_ops", "運用")):
            p = vars_[key].get().strip()
            if p and os.path.exists(p):
                defect_files[team] = p
        try:
            out = generate(
                folder_paths=[folder],
                output_path=vars_["output"].get().strip() or "pb_curve.xlsx",
                include_subfolders=sub.get(),
                defect_files=defect_files or None,
                pivot_date=_parse_cli_date(vars_["pivot"].get().strip() or None),
                start_date=_parse_cli_date(vars_["start"].get().strip() or None),
                end_date=_parse_cli_date(vars_["end"].get().strip() or None),
                b_final_rate=float(vars_["fr"].get()),
                b_lower_rate=float(vars_["lr"].get()),
                b_upper_rate=float(vars_["ur"].get()),
            )
            messagebox.showinfo("完了", f"出力しました:\n{os.path.abspath(out)}")
        except Exception as e:
            messagebox.showerror("エラー", f"生成に失敗しました:\n{e}")

    tk.Button(root, text="生成", width=16, command=on_run).grid(
        row=r, column=1, sticky="w", padx=6, pady=12)

    root.mainloop()


def main():
    # 引数なし起動（EXEダブルクリック等）はGUIを表示する
    if len(sys.argv) <= 1:
        run_gui()
        return

    parser = argparse.ArgumentParser(description="PB曲線（信頼度成長曲線）生成スクリプト")
    parser.add_argument("folder", nargs="*", help="対象フォルダパス（複数指定可）")
    parser.add_argument("-o", "--output", default="./output/pb_curve.xlsx",
                        help="出力ファイルパス（既定: ./output/pb_curve.xlsx）")
    parser.add_argument("--no-subfolders", action="store_true", help="サブフォルダを再帰探索しない")
    parser.add_argument("--pivot-date", help="基準日 YYYY-MM-DD（既定: 前営業日）")
    parser.add_argument("--start-date", help="開始日 YYYY-MM-DD（既定: データ最小日）")
    parser.add_argument("--end-date", help="終了日 YYYY-MM-DD（既定: データ最大日）")
    parser.add_argument("--b-final-rate", type=float, default=DEFAULT_B_FINAL_RATE,
                        help=f"B系最終計画係数（既定: {DEFAULT_B_FINAL_RATE}）")
    parser.add_argument("--b-lower-rate", type=float, default=DEFAULT_B_LOWER_RATE,
                        help=f"B系目標帯下限係数（既定: {DEFAULT_B_LOWER_RATE}）")
    parser.add_argument("--b-upper-rate", type=float, default=DEFAULT_B_UPPER_RATE,
                        help=f"B系目標帯上限係数（既定: {DEFAULT_B_UPPER_RATE}）")
    parser.add_argument("--forecast-mult", type=float, default=None,
                        help="予測倍率（既定: 実績から自動算出）")
    parser.add_argument("--total-case", type=int, default=None,
                        help="テストケース総数（既定: 収集件数）")
    parser.add_argument("--defect-online", help="欠陥一覧ファイル（オンライン）")
    parser.add_argument("--defect-batch", help="欠陥一覧ファイル（バッチ）")
    parser.add_argument("--defect-infra", help="欠陥一覧ファイル（基盤）")
    parser.add_argument("--defect-ops", help="欠陥一覧ファイル（運用）")
    args = parser.parse_args()

    if not args.folder:
        parser.print_help()
        print("\nエラー: 対象フォルダを1つ以上指定してください。")
        sys.exit(1)

    defect_files = {}
    for key, team in (("defect_online", "オンライン"), ("defect_batch", "バッチ"),
                      ("defect_infra", "基盤"), ("defect_ops", "運用")):
        path = getattr(args, key)
        if path and os.path.exists(path):
            defect_files[team] = path

    try:
        generate(
            folder_paths=args.folder,
            output_path=args.output,
            include_subfolders=not args.no_subfolders,
            defect_files=defect_files or None,
            pivot_date=_parse_cli_date(args.pivot_date),
            start_date=_parse_cli_date(args.start_date),
            end_date=_parse_cli_date(args.end_date),
            b_final_rate=args.b_final_rate,
            b_lower_rate=args.b_lower_rate,
            b_upper_rate=args.b_upper_rate,
            forecast_mult=args.forecast_mult,
            total_case=args.total_case,
        )
    except Exception as e:
        print(f"\nエラー: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
