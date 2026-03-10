"""
テスト進捗集計ツール - 検証用データ生成スクリプト

6ヶ月のプロジェクト（2025/12〜2026/05）を想定した
各種機能の確認が網羅できるテストデータを生成します。

生成されるデータ:
- 4チーム（オンライン、バッチ、基盤、運用）+ その他
- 各チームに複数のテストケースファイル
- 様々な進捗状況（完了、遅延、予定、未着手）
- 日付パターン（過去、当日、未来）
- サブフォルダ構造
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os
from datetime import datetime, timedelta
import random

# 設定
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "input")
SHEET_PREFIX = "ITB"
DATA_START_ROW = 19
COL_TEST_ID = 3  # C列
COL_JISSHI_YOTEI = 17  # Q列
COL_JISSHI_JISSEKI = 18  # R列
COL_KENSHO_YOTEI = 19  # S列
COL_KENSHO_JISSEKI = 20  # T列

# プロジェクト期間
PROJECT_START = datetime(2025, 12, 1)
PROJECT_END = datetime(2026, 5, 31)
TODAY = datetime(2026, 3, 5)  # 基準日

# チーム設定
TEAMS = {
    "オンライン": {
        "pattern": "-O-",
        "files": [
            {"name": "ITB-O-001_ログイン認証", "sheets": ["ITB-001_ログイン", "ITB-002_ログアウト", "ITB-003_パスワード変更"], "cases_per_sheet": 15},
            {"name": "ITB-O-002_ユーザー管理", "sheets": ["ITB-001_ユーザー登録", "ITB-002_ユーザー編集", "ITB-003_ユーザー削除"], "cases_per_sheet": 12},
            {"name": "ITB-O-003_検索機能", "sheets": ["ITB-001_商品検索", "ITB-002_注文検索"], "cases_per_sheet": 20},
            {"name": "ITB-O-004_カート機能", "sheets": ["ITB-001_カート追加", "ITB-002_カート編集", "ITB-003_カート削除"], "cases_per_sheet": 10},
        ],
        "subdir": None,
    },
    "バッチ": {
        "pattern": "-B-",
        "files": [
            {"name": "ITB-B-001_日次バッチ", "sheets": ["ITB-001_売上集計", "ITB-002_在庫更新", "ITB-003_レポート生成"], "cases_per_sheet": 8},
            {"name": "ITB-B-002_月次バッチ", "sheets": ["ITB-001_月次締め", "ITB-002_請求書発行"], "cases_per_sheet": 15},
            {"name": "ITB-B-003_データ連携", "sheets": ["ITB-001_外部システム連携", "ITB-002_データ同期"], "cases_per_sheet": 10},
        ],
        "subdir": None,
    },
    "基盤": {
        "pattern": "-I-",
        "files": [
            {"name": "ITB-I-001_認証基盤", "sheets": ["ITB-001_SSO", "ITB-002_トークン管理", "ITB-003_権限制御"], "cases_per_sheet": 12},
            {"name": "ITB-I-002_ログ基盤", "sheets": ["ITB-001_アクセスログ", "ITB-002_エラーログ", "ITB-003_監査ログ"], "cases_per_sheet": 8},
        ],
        "subdir": "基盤チーム",
    },
    "運用": {
        "pattern": "-U-",
        "files": [
            {"name": "ITB-U-001_監視機能", "sheets": ["ITB-001_死活監視", "ITB-002_性能監視", "ITB-003_アラート通知"], "cases_per_sheet": 10},
            {"name": "ITB-U-002_運用ツール", "sheets": ["ITB-001_バックアップ", "ITB-002_リストア", "ITB-003_メンテナンス"], "cases_per_sheet": 6},
        ],
        "subdir": "運用チーム",
    },
    "その他": {
        "pattern": "",  # パターンなし
        "files": [
            {"name": "ITB_共通機能テスト", "sheets": ["ITB-001_帳票出力", "ITB-002_メール送信"], "cases_per_sheet": 8},
            {"name": "ITB_外部連携テスト", "sheets": ["ITB-001_API連携"], "cases_per_sheet": 5},
        ],
        "subdir": "その他",
    },
}

# 進捗パターン
def generate_progress_pattern(test_id_num, total_cases, today):
    """テストケースの進捗パターンを生成"""
    progress_ratio = test_id_num / total_cases

    # プロジェクト進捗に基づく予定日を計算
    project_days = (PROJECT_END - PROJECT_START).days
    base_offset = int(project_days * progress_ratio)
    jisshi_yotei = PROJECT_START + timedelta(days=base_offset + random.randint(-5, 5))
    kensho_yotei = jisshi_yotei + timedelta(days=random.randint(1, 5))

    # 進捗状況を決定
    jisshi_jisseki = None
    kensho_jisseki = None

    if jisshi_yotei <= today:
        # 予定日が過去の場合
        if random.random() < 0.85:  # 85%は実施完了
            jisshi_jisseki = jisshi_yotei + timedelta(days=random.randint(-2, 3))
            if kensho_yotei <= today:
                if random.random() < 0.80:  # 80%は検証完了
                    kensho_jisseki = kensho_yotei + timedelta(days=random.randint(-1, 2))
                # 20%は検証遅延
            # 検証予定が未来の場合は検証未実施
        # 15%は実施遅延（実績なし）
    # 予定が未来の場合は未着手

    return jisshi_yotei, jisshi_jisseki, kensho_yotei, kensho_jisseki


def create_test_file(filepath, sheets_config, cases_per_sheet):
    """テストケースファイルを作成"""
    wb = openpyxl.Workbook()

    # 最初のシートを削除用にマーク
    first_sheet = True

    for sheet_name in sheets_config:
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        # ヘッダー行（18行目）
        header_row = 18
        headers = {
            COL_TEST_ID: "テストID",
            COL_JISSHI_YOTEI: "実施予定",
            COL_JISSHI_JISSEKI: "実施実績",
            COL_KENSHO_YOTEI: "検証予定",
            COL_KENSHO_JISSEKI: "検証実績",
        }

        for col, header in headers.items():
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # テストケースデータ
        for i in range(cases_per_sheet):
            row = DATA_START_ROW + i
            test_id = f"{sheet_name}-{i+1:03d}"

            jisshi_yotei, jisshi_jisseki, kensho_yotei, kensho_jisseki = generate_progress_pattern(
                i + 1, cases_per_sheet, TODAY
            )

            ws.cell(row=row, column=COL_TEST_ID, value=test_id)
            ws.cell(row=row, column=COL_JISSHI_YOTEI, value=jisshi_yotei)
            ws.cell(row=row, column=COL_JISSHI_JISSEKI, value=jisshi_jisseki)
            ws.cell(row=row, column=COL_KENSHO_YOTEI, value=kensho_yotei)
            ws.cell(row=row, column=COL_KENSHO_JISSEKI, value=kensho_jisseki)

            # 日付書式
            for col in [COL_JISSHI_YOTEI, COL_JISSHI_JISSEKI, COL_KENSHO_YOTEI, COL_KENSHO_JISSEKI]:
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.number_format = "YYYY/MM/DD"

    wb.save(filepath)
    print(f"  ✅ {os.path.basename(filepath)} ({len(sheets_config)}シート, 各{cases_per_sheet}件)")


def create_defect_detail_sheet(wb, team_name, num_records=25):
    """テスト欠陥一覧シートを既存ワークブックに追加

    Args:
        wb: openpyxl Workbook
        team_name: チーム名
        num_records: レコード数
    """
    ws = wb.create_sheet("テスト欠陥一覧")

    # 業務機能分類（全チーム共通）
    functions = [
        "01_委託者登録", "02_受付", "03_請求", "04_欠陥・返戻", "05_清算",
        "06_受入準備", "07_口振契約受付", "08_事務支(変更通知)", "09_事務支(その他)",
        "10_共通", "20_移行", "30_運用", "40_基盤",
    ]

    statuses = ["01:未着手", "02:調査中", "03:対応中", "04:検証中", "05:完了", "98:保留", "99:対応無し"]
    urgencies = ["高", "中", "低"]
    impacts = ["高", "中", "低"]
    root_causes = [
        "情報共有不足", "業務/仕様理解不足", "技術力不足", "影響範囲調査不足",
        "考慮不足", "注意不足", "プロセス不備", "未知", "外的要因", "非欠陥",
    ]
    embedded_phases = ["RD", "ED", "ID", "PD", "その他", "非欠陥"]
    detect_phases = ["CT", "ITa", "ITb", "ST", "非欠陥"]

    # ヘッダー行 (8行目)
    header_row = 8
    headers = {
        1: "欠陥ID",           # A
        2: "対応状況",          # B
        3: "件名",             # C
        4: "発見日",           # D
        7: "業務機能分類",      # G
        13: "緊急度",          # M
        14: "影響度",          # N
        15: "調査予定日",       # O
        16: "調査完了日",       # P
        20: "欠陥原因（深層）",  # T
        21: "欠陥埋込フェーズ",  # U
        22: "検出すべきフェーズ", # V
        30: "対応予定日",       # AD
        31: "対応日",          # AE
        33: "横展開有無",       # AG
        34: "横展開先",         # AH
        35: "横展開完了予定日",  # AI
        36: "横展開完了日",     # AJ
        37: "リリース予定日",    # AK
        38: "リリース日",       # AL
        39: "検証日",          # AM
        42: "集計フラグ",       # AP
    }

    for col, header in headers.items():
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # データ行 (9行目～)
    data_start = 9
    for i in range(num_records):
        row = data_start + i

        # 対応状況を決定（完了多め、全ステータス網羅）
        if i < num_records * 0.25:
            status = "05:完了"
        elif i < num_records * 0.40:
            status = "04:検証中"
        elif i < num_records * 0.55:
            status = "03:対応中"
        elif i < num_records * 0.65:
            status = "02:調査中"
        elif i < num_records * 0.75:
            status = "01:未着手"
        elif i < num_records * 0.85:
            status = "98:保留"
        else:
            status = "99:対応無し"

        # 集計フラグ
        count_flag = 0 if status == "99:対応無し" else 1

        # 発見日: プロジェクト期間内にランダム
        days_offset = random.randint(0, (TODAY - PROJECT_START).days)
        detected_date = PROJECT_START + timedelta(days=days_offset)

        # 調査予定日: 発見日から1-5日後
        investigate_plan = detected_date + timedelta(days=random.randint(1, 5))
        # 調査完了日
        investigate_done = None
        if status in ["03:対応中", "04:検証中", "05:完了"]:
            investigate_done = investigate_plan + timedelta(days=random.randint(-1, 3))

        # 対応予定日: 調査予定日から3-10日後
        fix_plan = investigate_plan + timedelta(days=random.randint(3, 10))
        # 対応日
        fix_done = None
        if status in ["04:検証中", "05:完了"]:
            fix_done = fix_plan + timedelta(days=random.randint(-2, 5))

        # 横展開
        has_lateral = random.random() < 0.3  # 30%が横展開あり
        lateral_exists = "有" if has_lateral else "無"
        lateral_target = random.choice(functions) if has_lateral else ""
        lateral_plan = None
        lateral_done = None
        if has_lateral:
            lateral_plan = fix_plan + timedelta(days=random.randint(3, 7)) if fix_plan else None
            if status in ["05:完了"] and random.random() < 0.7:
                lateral_done = lateral_plan + timedelta(days=random.randint(0, 3)) if lateral_plan else None

        # リリース予定日・リリース日
        release_plan = fix_plan + timedelta(days=random.randint(5, 14)) if fix_plan else None
        release_done = None
        if status in ["05:完了"]:
            release_done = release_plan + timedelta(days=random.randint(-1, 3)) if release_plan else None

        # 検証日
        verify_date = None
        if status in ["05:完了"]:
            verify_date = (fix_done or fix_plan) + timedelta(days=random.randint(1, 5))

        # 遅延パターン: 一部のレコードで調査予定超過/対応予定超過を作る
        if i % 7 == 0 and status in ["02:調査中"]:
            investigate_plan = TODAY - timedelta(days=random.randint(3, 10))
            investigate_done = None

        if i % 5 == 0 and status in ["03:対応中"]:
            fix_plan = TODAY - timedelta(days=random.randint(3, 10))
            fix_done = None

        # 滞留パターン: 発見日が古いのに未完了
        if i % 9 == 0 and status in ["02:調査中", "03:対応中"]:
            detected_date = TODAY - timedelta(days=random.randint(10, 30))

        # 欠陥ID
        defect_id = f"DEF-{team_name[0]}-{i+1:04d}"

        # セル設定
        ws.cell(row=row, column=1, value=defect_id)
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=f"欠陥{i+1}: {random.choice(functions)}の不具合")
        ws.cell(row=row, column=4, value=detected_date)
        ws.cell(row=row, column=7, value=random.choice(functions))
        ws.cell(row=row, column=13, value=random.choice(urgencies))
        ws.cell(row=row, column=14, value=random.choice(impacts))
        ws.cell(row=row, column=15, value=investigate_plan)
        ws.cell(row=row, column=16, value=investigate_done)
        ws.cell(row=row, column=20, value=random.choice(root_causes))
        ws.cell(row=row, column=21, value=random.choice(embedded_phases))
        ws.cell(row=row, column=22, value=random.choice(detect_phases))
        ws.cell(row=row, column=29, value=fix_plan)
        ws.cell(row=row, column=30, value=fix_done)
        ws.cell(row=row, column=32, value=lateral_exists)
        ws.cell(row=row, column=33, value=lateral_target)
        ws.cell(row=row, column=34, value=lateral_plan)
        ws.cell(row=row, column=35, value=lateral_done)
        ws.cell(row=row, column=37, value=release_plan)
        ws.cell(row=row, column=38, value=release_done)
        ws.cell(row=row, column=39, value=verify_date)
        ws.cell(row=row, column=42, value=count_flag)

        # 日付書式
        date_cols = [4, 15, 16, 30, 31, 35, 36, 37, 38, 39]
        for col in date_cols:
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.number_format = "YYYY/MM/DD"

    print(f"    テスト欠陥一覧: {num_records}件")


def main():
    """メイン処理"""
    print("=" * 60)
    print("  テスト進捗集計ツール - 検証用データ生成")
    print("=" * 60)
    print(f"\n  プロジェクト期間: {PROJECT_START.strftime('%Y/%m/%d')} 〜 {PROJECT_END.strftime('%Y/%m/%d')}")
    print(f"  基準日: {TODAY.strftime('%Y/%m/%d')}")
    print(f"  出力先: {OUTPUT_DIR}\n")

    # 既存ファイルを削除
    if os.path.exists(OUTPUT_DIR):
        for root, dirs, files in os.walk(OUTPUT_DIR, topdown=False):
            for name in files:
                if name.endswith(('.xlsx', '.xlsm')) and not name.startswith('~$'):
                    os.remove(os.path.join(root, name))
            for name in dirs:
                try:
                    os.rmdir(os.path.join(root, name))
                except OSError:
                    pass

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    total_files = 0
    total_sheets = 0
    total_cases = 0

    for team_name, team_config in TEAMS.items():
        print(f"\n[{team_name}チーム]")

        # サブディレクトリ
        if team_config["subdir"]:
            team_dir = os.path.join(OUTPUT_DIR, team_config["subdir"])
            os.makedirs(team_dir, exist_ok=True)
        else:
            team_dir = OUTPUT_DIR

        for file_config in team_config["files"]:
            filename = f"{file_config['name']}.xlsx"
            filepath = os.path.join(team_dir, filename)

            create_test_file(
                filepath,
                file_config["sheets"],
                file_config["cases_per_sheet"]
            )

            total_files += 1
            total_sheets += len(file_config["sheets"])
            total_cases += len(file_config["sheets"]) * file_config["cases_per_sheet"]

    # 特殊ケース: xlsmファイル
    print("\n[特殊ケース]")
    xlsm_path = os.path.join(OUTPUT_DIR, "ITB-O-005_マクロ付きテスト.xlsm")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ITB-001_マクロテスト"

    # ヘッダー
    for col, header in [(COL_TEST_ID, "テストID"), (COL_JISSHI_YOTEI, "実施予定"),
                        (COL_JISSHI_JISSEKI, "実施実績"), (COL_KENSHO_YOTEI, "検証予定"),
                        (COL_KENSHO_JISSEKI, "検証実績")]:
        ws.cell(row=18, column=col, value=header).font = Font(bold=True)

    # データ
    for i in range(5):
        row = DATA_START_ROW + i
        ws.cell(row=row, column=COL_TEST_ID, value=f"ITB-001_マクロテスト-{i+1:03d}")
        ws.cell(row=row, column=COL_JISSHI_YOTEI, value=TODAY - timedelta(days=10-i*2))
        ws.cell(row=row, column=COL_JISSHI_JISSEKI, value=TODAY - timedelta(days=8-i*2) if i < 3 else None)
        ws.cell(row=row, column=COL_KENSHO_YOTEI, value=TODAY - timedelta(days=5-i*2))
        ws.cell(row=row, column=COL_KENSHO_JISSEKI, value=TODAY - timedelta(days=3-i*2) if i < 2 else None)
        for col in [COL_JISSHI_YOTEI, COL_JISSHI_JISSEKI, COL_KENSHO_YOTEI, COL_KENSHO_JISSEKI]:
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.number_format = "YYYY/MM/DD"

    wb.save(xlsm_path)
    print(f"  ✅ ITB-O-005_マクロ付きテスト.xlsm (xlsm形式)")
    total_files += 1
    total_sheets += 1
    total_cases += 5

    # 対象外ファイル（ITBで始まらないシート）
    non_target_path = os.path.join(OUTPUT_DIR, "参考資料_設計書.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "設計書"  # ITBで始まらない
    ws['A1'] = "これは参考資料です（集計対象外）"
    wb.save(non_target_path)
    print(f"  ✅ 参考資料_設計書.xlsx (対象外ファイル)")

    # 欠陥ファイルの生成（欠陥発見・対応推移集計表 + テスト欠陥一覧）
    defect_dir = os.path.join(OUTPUT_DIR, "defects")
    os.makedirs(defect_dir, exist_ok=True)
    print("\n[欠陥ファイル生成]")
    defect_team_map = {
        "オンライン": "欠陥一覧_オンライン.xlsx",
        "バッチ": "欠陥一覧_バッチ.xlsx",
        "基盤": "欠陥一覧_基盤.xlsx",
        "運用": "欠陥一覧_運用.xlsx",
    }
    defect_record_counts = {
        "オンライン": 30,
        "バッチ": 20,
        "基盤": 15,
        "運用": 18,
    }
    for team_name, filename in defect_team_map.items():
        filepath = os.path.join(defect_dir, filename)
        wb = openpyxl.Workbook()

        # 欠陥発見・対応推移集計表シート
        ws = wb.active
        ws.title = "欠陥発見・対応推移集計表"

        # ヘッダー (10行目)
        header_row = 10
        headers_defect = {
            2: "No.",      # B列
            3: "日付",     # C列
            4: "検出",     # D列
            5: "対応",     # E列
            6: "累積検出", # F列
            7: "累積対応", # G列
            8: "累積未対応", # H列
        }
        for col, header in headers_defect.items():
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # データ (11行目～): プロジェクト期間の営業日分
        data_row = 11
        cum_detected = 0
        cum_resolved = 0
        current_date = PROJECT_START
        no = 1
        while current_date <= TODAY:
            if current_date.weekday() < 5:  # 平日のみ
                detected = random.choice([0, 0, 0, 1, 1, 2]) if random.random() < 0.6 else 0
                resolved = random.choice([0, 0, 1, 1]) if cum_detected > cum_resolved else 0
                cum_detected += detected
                cum_resolved += resolved

                ws.cell(row=data_row, column=2, value=no)
                ws.cell(row=data_row, column=3, value=current_date)
                ws.cell(row=data_row, column=3).number_format = "YYYY/MM/DD"
                ws.cell(row=data_row, column=4, value=detected)
                ws.cell(row=data_row, column=5, value=resolved)
                ws.cell(row=data_row, column=6, value=cum_detected)
                ws.cell(row=data_row, column=7, value=cum_resolved)
                ws.cell(row=data_row, column=8, value=cum_detected - cum_resolved)

                data_row += 1
                no += 1
            current_date += timedelta(days=1)

        # テスト欠陥一覧シートを追加
        num_records = defect_record_counts[team_name]
        create_defect_detail_sheet(wb, team_name, num_records)

        wb.save(filepath)
        print(f"  ✅ {filename} (推移集計: {no-1}日分, 欠陥詳細: {num_records}件)")

    print("\n" + "=" * 60)
    print(f"  生成完了!")
    print(f"  - ファイル数: {total_files}")
    print(f"  - シート数: {total_sheets}")
    print(f"  - テストケース数: {total_cases}")
    print("=" * 60)


if __name__ == "__main__":
    main()
