"""
============================================================
  テスト予定・実績 集計スクリプト v4
============================================================
  概要：
    指定フォルダ内の全Excelファイルから
    「ITB」で始まるシートを対象に、
    実施者・検証者の予定/実績日付を読み取り、
    明細＋進捗サマリー＋祝日の3シート構成でExcelに出力します。

  出力構成：
    [進捗サマリー]   日付×予定/実績の件数集計（明細参照式+累計+進捗判定）
    [明細シート]     ファイル/シート/テストID単位の全レコード（テーブル形式）
    [祝日マスタ]     日本の祝日を管理（営業日判定用）

  主な機能：
    - ウィザード形式の使いやすいUI
    - チーム名自動識別（-O-:オンライン, -B-:バッチ, -I-:基盤, -U-:運用）
    - サブフォルダ含む再帰的なファイル収集
    - 差分更新（前回集計済みファイルはスキップ）
    - 日付別サマリーは明細シートを関数で参照
    - 進捗判定列（順調/遅延/完了/予定）
    - 条件付き書式による進捗の視覚化
    - 累計列の自動計算
    - 営業日/非営業日の識別
    - 基準日セルによる進捗基準の設定

  使い方：
    1. Python で実行: python aggregate_test_results.py
    ※ EXE化: pyinstaller --onefile --windowed aggregate_test_results.py

  EXE化時の注意：
    - --windowedオプションでコンソール非表示
    - tkinterは標準ライブラリなのでそのまま使用可能
============================================================
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule, CellIsRule, DataBarRule
import os
import sys
import argparse
from collections import defaultdict
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json

# ===================================================================
#  ===設定=== ここを環境に合わせて変更してください
# ===================================================================

# --- シートのフィルタ条件 ---
SHEET_PREFIX = "ITB"

# --- データ位置の設定 ---
COL_TEST_ID = 3  # C列
COL_JISSHI_YOTEI   = 17  # Q列: 実施者 予定
COL_JISSHI_JISSEKI  = 18  # R列: 実施者 実績
COL_KENSHO_YOTEI   = 19  # S列: 検証者 予定
COL_KENSHO_JISSEKI  = 20  # T列: 検証者 実績
DATA_START_ROW = 19

# --- チーム識別パターン ---
TEAM_PATTERNS = {
    "-O-": "オンライン",
    "-B-": "バッチ",
    "-I-": "基盤",
    "-U-": "運用",
}

# --- 日本の祝日（デフォルト: 2024-2030年） ---
# 辞書形式: {日付: 祝日名}
DEFAULT_HOLIDAYS = {
    # 2024年
    "2024/01/01": "元日",
    "2024/01/02": "年始休暇",
    "2024/01/03": "年始休暇",
    "2024/01/08": "成人の日",
    "2024/02/11": "建国記念の日",
    "2024/02/12": "振替休日",
    "2024/02/23": "天皇誕生日",
    "2024/03/20": "春分の日",
    "2024/04/29": "昭和の日",
    "2024/05/03": "憲法記念日",
    "2024/05/04": "みどりの日",
    "2024/05/05": "こどもの日",
    "2024/05/06": "振替休日",
    "2024/07/15": "海の日",
    "2024/08/11": "山の日",
    "2024/08/12": "振替休日",
    "2024/09/16": "敬老の日",
    "2024/09/22": "秋分の日",
    "2024/09/23": "振替休日",
    "2024/10/14": "スポーツの日",
    "2024/11/03": "文化の日",
    "2024/11/04": "振替休日",
    "2024/11/23": "勤労感謝の日",
    "2024/12/30": "年末休暇",
    "2024/12/31": "年末休暇",
    # 2025年
    "2025/01/01": "元日",
    "2025/01/02": "年始休暇",
    "2025/01/03": "年始休暇",
    "2025/01/13": "成人の日",
    "2025/02/11": "建国記念の日",
    "2025/02/23": "天皇誕生日",
    "2025/02/24": "振替休日",
    "2025/03/20": "春分の日",
    "2025/04/29": "昭和の日",
    "2025/05/03": "憲法記念日",
    "2025/05/04": "みどりの日",
    "2025/05/05": "こどもの日",
    "2025/05/06": "振替休日",
    "2025/07/21": "海の日",
    "2025/08/11": "山の日",
    "2025/09/15": "敬老の日",
    "2025/09/23": "秋分の日",
    "2025/10/13": "スポーツの日",
    "2025/11/03": "文化の日",
    "2025/11/23": "勤労感謝の日",
    "2025/11/24": "振替休日",
    "2025/12/30": "年末休暇",
    "2025/12/31": "年末休暇",
    # 2026年
    "2026/01/01": "元日",
    "2026/01/02": "年始休暇",
    "2026/01/03": "年始休暇",
    "2026/01/12": "成人の日",
    "2026/02/11": "建国記念の日",
    "2026/02/23": "天皇誕生日",
    "2026/03/20": "春分の日",
    "2026/04/29": "昭和の日",
    "2026/05/03": "憲法記念日",
    "2026/05/04": "みどりの日",
    "2026/05/05": "こどもの日",
    "2026/05/06": "振替休日",
    "2026/07/20": "海の日",
    "2026/08/11": "山の日",
    "2026/09/21": "敬老の日",
    "2026/09/22": "国民の休日",
    "2026/09/23": "秋分の日",
    "2026/10/12": "スポーツの日",
    "2026/11/03": "文化の日",
    "2026/11/23": "勤労感謝の日",
    "2026/12/30": "年末休暇",
    "2026/12/31": "年末休暇",
    # 2027年
    "2027/01/01": "元日",
    "2027/01/02": "年始休暇",
    "2027/01/03": "年始休暇",
    "2027/01/11": "成人の日",
    "2027/02/11": "建国記念の日",
    "2027/02/23": "天皇誕生日",
    "2027/03/21": "春分の日",
    "2027/03/22": "振替休日",
    "2027/04/29": "昭和の日",
    "2027/05/03": "憲法記念日",
    "2027/05/04": "みどりの日",
    "2027/05/05": "こどもの日",
    "2027/07/19": "海の日",
    "2027/08/11": "山の日",
    "2027/09/20": "敬老の日",
    "2027/09/23": "秋分の日",
    "2027/10/11": "スポーツの日",
    "2027/11/03": "文化の日",
    "2027/11/23": "勤労感謝の日",
    "2027/12/30": "年末休暇",
    "2027/12/31": "年末休暇",
    # 2028年
    "2028/01/01": "元日",
    "2028/01/02": "年始休暇",
    "2028/01/03": "年始休暇",
    "2028/01/10": "成人の日",
    "2028/02/11": "建国記念の日",
    "2028/02/23": "天皇誕生日",
    "2028/03/20": "春分の日",
    "2028/04/29": "昭和の日",
    "2028/05/03": "憲法記念日",
    "2028/05/04": "みどりの日",
    "2028/05/05": "こどもの日",
    "2028/07/17": "海の日",
    "2028/08/11": "山の日",
    "2028/09/18": "敬老の日",
    "2028/09/22": "秋分の日",
    "2028/10/09": "スポーツの日",
    "2028/11/03": "文化の日",
    "2028/11/23": "勤労感謝の日",
    "2028/12/30": "年末休暇",
    "2028/12/31": "年末休暇",
    # 2029年
    "2029/01/01": "元日",
    "2029/01/02": "年始休暇",
    "2029/01/03": "年始休暇",
    "2029/01/08": "成人の日",
    "2029/02/11": "建国記念の日",
    "2029/02/12": "振替休日",
    "2029/02/23": "天皇誕生日",
    "2029/03/20": "春分の日",
    "2029/04/29": "昭和の日",
    "2029/04/30": "振替休日",
    "2029/05/03": "憲法記念日",
    "2029/05/04": "みどりの日",
    "2029/05/05": "こどもの日",
    "2029/07/16": "海の日",
    "2029/08/11": "山の日",
    "2029/09/17": "敬老の日",
    "2029/09/23": "秋分の日",
    "2029/09/24": "振替休日",
    "2029/10/08": "スポーツの日",
    "2029/11/03": "文化の日",
    "2029/11/23": "勤労感謝の日",
    "2029/12/30": "年末休暇",
    "2029/12/31": "年末休暇",
    # 2030年
    "2030/01/01": "元日",
    "2030/01/02": "年始休暇",
    "2030/01/03": "年始休暇",
    "2030/01/14": "成人の日",
    "2030/02/11": "建国記念の日",
    "2030/02/23": "天皇誕生日",
    "2030/03/20": "春分の日",
    "2030/04/29": "昭和の日",
    "2030/05/03": "憲法記念日",
    "2030/05/04": "みどりの日",
    "2030/05/05": "こどもの日",
    "2030/05/06": "振替休日",
    "2030/07/15": "海の日",
    "2030/08/11": "山の日",
    "2030/08/12": "振替休日",
    "2030/09/16": "敬老の日",
    "2030/09/23": "秋分の日",
    "2030/10/14": "スポーツの日",
    "2030/11/03": "文化の日",
    "2030/11/04": "振替休日",
    "2030/11/23": "勤労感謝の日",
    "2030/12/30": "年末休暇",
    "2030/12/31": "年末休暇",
}

# ===================================================================
#  スタイル定義（4色システム: 緑=完了, 黄=進行中, 赤=遅延, グレー=対象外）
# ===================================================================

# --- 基本フォント・配置 ---
TITLE_FONT = Font(name="游ゴシック", size=14, bold=True, color="333333")
TITLE_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

HEADER_FONT = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="505050", end_color="505050", fill_type="solid")  # ダークグレー
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="游ゴシック", size=10)
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
DATA_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

THICK_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="medium"), bottom=Side(style="medium"),
)

# --- 4色システム（信号機モデル） ---
# 緑: 完了・安全
COMPLETE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
COMPLETE_FONT = Font(name="游ゴシック", size=10, bold=True, color="006100")

# 黄: 注意・進行中（予定通り）
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
WARNING_FONT = Font(name="游ゴシック", size=10, bold=True, color="9C5700")

# 赤: 危険・遅延
DANGER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
DANGER_FONT = Font(name="游ゴシック", size=10, bold=True, color="9C0006")

# グレー: 対象外・未来・非営業日
NEUTRAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
NEUTRAL_FONT = Font(name="游ゴシック", size=10, color="808080")

# --- 合計行（罫線で区切り、薄いグレー背景） ---
TOTAL_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

# --- 基準日セルのスタイル ---
REF_DATE_FONT = Font(name="游ゴシック", size=12, bold=True, color="FFFFFF")
REF_DATE_FILL = PatternFill(start_color="505050", end_color="505050", fill_type="solid")

# --- ダッシュボード用スタイル ---
DASHBOARD_TITLE_FONT = Font(name="游ゴシック", size=16, bold=True, color="333333")
DASHBOARD_SECTION_FONT = Font(name="游ゴシック", size=12, bold=True, color="505050")
DASHBOARD_VALUE_FONT = Font(name="游ゴシック", size=24, bold=True)
DASHBOARD_LABEL_FONT = Font(name="游ゴシック", size=10, color="666666")

# --- ダッシュボード配色（実施/検証テーマ） ---
# 実施（Implementation）テーマ: 青系
DASHBOARD_IMPL_SECTION_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
DASHBOARD_IMPL_SECTION_FONT = Font(name="游ゴシック", size=12, bold=True, color="FFFFFF")
DASHBOARD_IMPL_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DASHBOARD_IMPL_HEADER_FONT = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")

# 検証（Verification）テーマ: オレンジ系
DASHBOARD_VERIFY_SECTION_FILL = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
DASHBOARD_VERIFY_SECTION_FONT = Font(name="游ゴシック", size=12, bold=True, color="FFFFFF")
DASHBOARD_VERIFY_HEADER_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
DASHBOARD_VERIFY_HEADER_FONT = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")

# --- 順調ステータス（青系、指示書準拠） ---
OK_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
OK_FONT = Font(name="游ゴシック", size=10, bold=True, color="2F5597")

# --- 罫線スタイル ---
THIN_SOLID_SIDE = Side(style='thin', color='000000')
THIN_DOTTED_SIDE = Side(style='dotted', color='000000')
MEDIUM_SOLID_SIDE = Side(style='medium', color='000000')

# 列グループ用罫線（内部は点線）
def create_dotted_border(left_solid=False, right_solid=False):
    """列グループ用罫線を作成（内部は点線、境界は実線）"""
    return Border(
        left=THIN_SOLID_SIDE if left_solid else THIN_DOTTED_SIDE,
        right=THIN_SOLID_SIDE if right_solid else THIN_DOTTED_SIDE,
        top=THIN_SOLID_SIDE,
        bottom=THIN_SOLID_SIDE,
    )

# --- 進捗サマリーシート新配色 ---
SUMMARY_TITLE_BG = "1B3A5C"           # 行1: ダークネイビー
SUMMARY_SUBTITLE_BG = "E8EEF4"        # 行2: 薄いブルーグレー
SUMMARY_SUMMARY_ROW_BG = "D9E2F3"     # 行5: 合計行の薄いブルー

# カテゴリグループヘッダー（行3）
SUMMARY_GROUP_COMMON = "505050"       # 共通
SUMMARY_GROUP_IMPL = "2B5797"         # 実施
SUMMARY_GROUP_VERIFY = "2E7D32"       # 検証
SUMMARY_GROUP_TOTAL = "E65100"        # 合計

# サブヘッダー（行4）
SUMMARY_SUB_COMMON = "6D6D6D"         # 共通
SUMMARY_SUB_IMPL = "4472C4"           # 実施
SUMMARY_SUB_VERIFY = "548235"         # 検証
SUMMARY_SUB_TOTAL = "ED7D31"          # 合計

# ステータス条件付き書式（新配色）
STATUS_COLORS = {
    "遅延": {"bg": "FFE0E0", "fg": "D32F2F", "bold": True},
    "順調": {"bg": "E3F2FD", "fg": "1565C0", "bold": True},
    "完了": {"bg": "E8F5E9", "fg": "2E7D32", "bold": True},
    "予定": {"bg": "F5F5F5", "fg": "9E9E9E", "bold": False},
    "－": {"bg": "FAFAFA", "fg": "BDBDBD", "bold": False},
}

# 基準日ハイライト
BASEDATE_HIGHLIGHT_BG = "FFF9C4"
BASEDATE_HIGHLIGHT_FG = "E65100"

# データバー色
DATABAR_IMPL = "BDD7EE"
DATABAR_VERIFY = "C6EFCE"

# 二重線罫線
DOUBLE_SIDE = Side(style='double', color='000000')

# 列幅（pt÷7.5で文字幅に変換）
SUMMARY_COL_WIDTHS = {
    'A': 14.0, 'B': 4.0, 'C': 8.3, 'D': 8.0, 'E': 8.0,
    'F': 9.1, 'G': 10.4, 'H': 10.4, 'I': 12.8, 'J': 8.0,
    'K': 8.0, 'L': 8.0, 'M': 9.1, 'N': 10.4, 'O': 10.4,
    'P': 12.8, 'Q': 8.0, 'R': 9.6, 'S': 13.0,
}


# ===================================================================
#  ウィザードUI
# ===================================================================

class WizardApp(tk.Tk):
    """ウィザード形式のメインアプリケーション"""

    def __init__(self):
        super().__init__()

        self.title("テスト進捗集計ツール v4")
        self.geometry("600x560")
        self.resizable(False, False)

        # 常に最前面に表示
        self.attributes("-topmost", True)

        # 結果を格納する変数
        self.result = None
        self.folder_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.include_subfolders = tk.BooleanVar(value=True)
        self.update_mode = tk.StringVar(value="new")  # "new" or "update"

        # 週集計範囲（From/To）
        self.week_from = tk.StringVar()
        self.week_to = tk.StringVar()

        # 現在のステップ
        self.current_step = 1

        # メインフレーム
        self.main_frame = ttk.Frame(self, padding=20)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # ヘッダー
        self.header_frame = ttk.Frame(self.main_frame)
        self.header_frame.pack(fill=tk.X, pady=(0, 20))

        self.title_label = ttk.Label(
            self.header_frame,
            text="テスト進捗集計ツール",
            font=("游ゴシック", 16, "bold")
        )
        self.title_label.pack()

        self.step_label = ttk.Label(
            self.header_frame,
            text="ステップ 1/4: 対象フォルダ選択",
            font=("游ゴシック", 11)
        )
        self.step_label.pack(pady=(5, 0))

        # コンテンツフレーム
        self.content_frame = ttk.Frame(self.main_frame)
        self.content_frame.pack(fill=tk.BOTH, expand=True)

        # ボタンフレーム
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.pack(fill=tk.X, pady=(20, 0))

        self.back_btn = ttk.Button(self.button_frame, text="< 戻る", command=self.go_back)
        self.back_btn.pack(side=tk.LEFT)

        self.cancel_btn = ttk.Button(self.button_frame, text="キャンセル", command=self.cancel)
        self.cancel_btn.pack(side=tk.RIGHT, padx=(10, 0))

        self.next_btn = ttk.Button(self.button_frame, text="次へ >", command=self.go_next)
        self.next_btn.pack(side=tk.RIGHT)

        # 最初のステップを表示
        self.show_step(1)

        # ウィンドウを中央に配置
        self.center_window()

    def center_window(self):
        """ウィンドウを画面中央に配置"""
        self.update_idletasks()
        width = 600
        height = 560
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

    def clear_content(self):
        """コンテンツフレームをクリア"""
        for widget in self.content_frame.winfo_children():
            widget.destroy()

    def show_step(self, step):
        """指定されたステップを表示"""
        self.current_step = step
        self.clear_content()

        if step == 1:
            self.show_step1()
        elif step == 2:
            self.show_step2()
        elif step == 3:
            self.show_step3()
        elif step == 4:
            self.show_step4()

        # ボタンの状態を更新
        self.back_btn.config(state=tk.NORMAL if step > 1 else tk.DISABLED)
        self.next_btn.config(text="実行" if step == 4 else "次へ >")

    def show_step1(self):
        """ステップ1: 対象フォルダ選択"""
        self.step_label.config(text="ステップ 1/4: 対象フォルダ選択")

        # 説明
        desc = ttk.Label(
            self.content_frame,
            text="集計対象のExcelファイルが格納されているフォルダを選択してください。",
            wraplength=480
        )
        desc.pack(anchor=tk.W, pady=(0, 15))

        # フォルダ選択
        folder_frame = ttk.Frame(self.content_frame)
        folder_frame.pack(fill=tk.X, pady=10)

        ttk.Button(
            folder_frame,
            text="📁 フォルダを選択...",
            command=self.select_folder
        ).pack(side=tk.LEFT)

        self.folder_display = ttk.Label(
            folder_frame,
            text=self.folder_path.get() or "(未選択)",
            foreground="gray" if not self.folder_path.get() else "black"
        )
        self.folder_display.pack(side=tk.LEFT, padx=(10, 0))

        # サブフォルダオプション
        ttk.Checkbutton(
            self.content_frame,
            text="サブフォルダも含める（推奨）",
            variable=self.include_subfolders
        ).pack(anchor=tk.W, pady=(20, 0))

        # 説明追加
        note = ttk.Label(
            self.content_frame,
            text="※ サブフォルダを含めると、選択したフォルダ配下の全てのExcelファイルが対象になります。",
            foreground="gray",
            wraplength=480
        )
        note.pack(anchor=tk.W, pady=(5, 0))

    def show_step2(self):
        """ステップ2: 週集計範囲"""
        self.step_label.config(text="ステップ 2/4: 週集計範囲")

        # 説明
        desc = ttk.Label(
            self.content_frame,
            text="ダッシュボードの「週次」セクションで集計する週の範囲を指定してください。\n空欄の場合は週次集計を行いません。",
            wraplength=480
        )
        desc.pack(anchor=tk.W, pady=(0, 15))

        # 週範囲入力フレーム
        week_frame = ttk.LabelFrame(self.content_frame, text="週集計範囲（任意）", padding=15)
        week_frame.pack(fill=tk.X, pady=10)

        # From
        from_frame = ttk.Frame(week_frame)
        from_frame.pack(fill=tk.X, pady=5)
        ttk.Label(from_frame, text="From（開始日）:", width=15, anchor=tk.W).pack(side=tk.LEFT)
        self.week_from_entry = ttk.Entry(from_frame, textvariable=self.week_from, width=15)
        self.week_from_entry.pack(side=tk.LEFT)
        ttk.Label(from_frame, text="  例: 2025/01/06", foreground="gray").pack(side=tk.LEFT)

        # To
        to_frame = ttk.Frame(week_frame)
        to_frame.pack(fill=tk.X, pady=5)
        ttk.Label(to_frame, text="To（終了日）:", width=15, anchor=tk.W).pack(side=tk.LEFT)
        self.week_to_entry = ttk.Entry(to_frame, textvariable=self.week_to, width=15)
        self.week_to_entry.pack(side=tk.LEFT)
        ttk.Label(to_frame, text="  例: 2025/01/10", foreground="gray").pack(side=tk.LEFT)

        # 注意書き
        note = ttk.Label(
            self.content_frame,
            text="※ 日付は YYYY/MM/DD 形式で入力してください。\n※ 週次集計は指定した期間内の予定・実績を集計します。",
            foreground="gray",
            wraplength=480
        )
        note.pack(anchor=tk.W, pady=(15, 0))

    def show_step3(self):
        """ステップ3: 出力設定"""
        self.step_label.config(text="ステップ 3/4: 出力設定")

        # 説明
        desc = ttk.Label(
            self.content_frame,
            text="集計結果の出力方法を選択してください。",
            wraplength=480
        )
        desc.pack(anchor=tk.W, pady=(0, 15))

        # モード選択
        mode_frame = ttk.LabelFrame(self.content_frame, text="出力モード", padding=10)
        mode_frame.pack(fill=tk.X, pady=10)

        ttk.Radiobutton(
            mode_frame,
            text="新規作成",
            variable=self.update_mode,
            value="new"
        ).pack(anchor=tk.W)

        ttk.Label(
            mode_frame,
            text="   新しいExcelファイルを作成します",
            foreground="gray"
        ).pack(anchor=tk.W)

        ttk.Radiobutton(
            mode_frame,
            text="既存ファイルを更新（上書き）",
            variable=self.update_mode,
            value="update"
        ).pack(anchor=tk.W, pady=(10, 0))

        ttk.Label(
            mode_frame,
            text="   既存のExcelファイルを選択し、最新データで上書きします",
            foreground="gray"
        ).pack(anchor=tk.W)

        # ファイル選択
        file_frame = ttk.Frame(self.content_frame)
        file_frame.pack(fill=tk.X, pady=(20, 10))

        ttk.Button(
            file_frame,
            text="💾 保存先を選択...",
            command=self.select_output
        ).pack(side=tk.LEFT)

        self.output_display = ttk.Label(
            file_frame,
            text=self.output_path.get() or "(未選択)",
            foreground="gray" if not self.output_path.get() else "black"
        )
        self.output_display.pack(side=tk.LEFT, padx=(10, 0))

    def show_step4(self):
        """ステップ4: 確認"""
        self.step_label.config(text="ステップ 4/4: 確認")

        # 説明
        desc = ttk.Label(
            self.content_frame,
            text="以下の設定で集計を実行します。内容を確認してください。",
            wraplength=480
        )
        desc.pack(anchor=tk.W, pady=(0, 15))

        # 設定内容表示
        confirm_frame = ttk.LabelFrame(self.content_frame, text="設定内容", padding=15)
        confirm_frame.pack(fill=tk.X, pady=10)

        # フォルダ
        row1 = ttk.Frame(confirm_frame)
        row1.pack(fill=tk.X, pady=3)
        ttk.Label(row1, text="対象フォルダ:", width=15, anchor=tk.W).pack(side=tk.LEFT)
        ttk.Label(row1, text=self.folder_path.get(), wraplength=350).pack(side=tk.LEFT)

        # サブフォルダ
        row2 = ttk.Frame(confirm_frame)
        row2.pack(fill=tk.X, pady=3)
        ttk.Label(row2, text="サブフォルダ:", width=15, anchor=tk.W).pack(side=tk.LEFT)
        ttk.Label(row2, text="含める" if self.include_subfolders.get() else "含めない").pack(side=tk.LEFT)

        # 出力先
        row3 = ttk.Frame(confirm_frame)
        row3.pack(fill=tk.X, pady=3)
        ttk.Label(row3, text="出力先:", width=15, anchor=tk.W).pack(side=tk.LEFT)
        ttk.Label(row3, text=self.output_path.get(), wraplength=350).pack(side=tk.LEFT)

        # モード
        row4 = ttk.Frame(confirm_frame)
        row4.pack(fill=tk.X, pady=3)
        ttk.Label(row4, text="モード:", width=15, anchor=tk.W).pack(side=tk.LEFT)
        mode_text = "新規作成" if self.update_mode.get() == "new" else "既存ファイル更新"
        ttk.Label(row4, text=mode_text).pack(side=tk.LEFT)

        # 週集計範囲
        row5 = ttk.Frame(confirm_frame)
        row5.pack(fill=tk.X, pady=3)
        ttk.Label(row5, text="週集計範囲:", width=15, anchor=tk.W).pack(side=tk.LEFT)
        week_from = self.week_from.get()
        week_to = self.week_to.get()
        if week_from and week_to:
            week_text = f"{week_from} ～ {week_to}"
        else:
            week_text = "（指定なし）"
        ttk.Label(row5, text=week_text).pack(side=tk.LEFT)

        # 注意書き
        note = ttk.Label(
            self.content_frame,
            text="※ 「実行」をクリックすると集計処理を開始します。\n※ 前回集計済みのファイルは自動的にスキップされます。",
            foreground="gray",
            wraplength=480
        )
        note.pack(anchor=tk.W, pady=(15, 0))

    def select_folder(self):
        """フォルダ選択ダイアログ"""
        folder = filedialog.askdirectory(title="対象フォルダを選択")
        if folder:
            self.folder_path.set(folder)
            self.folder_display.config(text=folder, foreground="black")

    def select_output(self):
        """出力ファイル選択ダイアログ"""
        if self.update_mode.get() == "new":
            # 新規作成モード: ファイル名を入力して保存先を選択
            path = filedialog.asksaveasfilename(
                title="保存先を選択",
                defaultextension=".xlsx",
                filetypes=[("Excel ファイル", "*.xlsx")],
                initialfile="テスト進捗集計.xlsx",
                confirmoverwrite=False  # OS標準の上書き確認を無効化
            )
            if path:
                # 既存ファイルの場合は日本語で確認
                if os.path.exists(path):
                    confirm = messagebox.askyesno(
                        "上書き確認",
                        f"ファイルが既に存在します。\n\n{os.path.basename(path)}\n\n上書きしてもよろしいですか？"
                    )
                    if not confirm:
                        return
        else:
            # 更新モード: 既存ファイルを選択
            path = filedialog.askopenfilename(
                title="更新するファイルを選択",
                filetypes=[("Excel ファイル", "*.xlsx")]
            )
            if path:
                # 更新モードでも確認
                confirm = messagebox.askyesno(
                    "更新確認",
                    f"以下のファイルを最新データで上書きします。\n\n{os.path.basename(path)}\n\nよろしいですか？"
                )
                if not confirm:
                    return

        if path:
            self.output_path.set(path)
            self.output_display.config(text=path, foreground="black")

    def go_back(self):
        """前のステップへ"""
        if self.current_step > 1:
            self.show_step(self.current_step - 1)

    def go_next(self):
        """次のステップへ / 実行"""
        if self.current_step == 1:
            if not self.folder_path.get():
                messagebox.showwarning("入力エラー", "対象フォルダを選択してください。")
                return
            self.show_step(2)

        elif self.current_step == 2:
            # 週集計範囲のバリデーション（入力がある場合のみ）
            week_from = self.week_from.get().strip()
            week_to = self.week_to.get().strip()
            if week_from or week_to:
                # どちらか一方だけ入力されている場合はエラー
                if not week_from or not week_to:
                    messagebox.showwarning("入力エラー", "週集計範囲を指定する場合は、開始日と終了日の両方を入力してください。")
                    return
                # 日付形式のバリデーション
                try:
                    from_date = datetime.strptime(week_from, "%Y/%m/%d")
                    to_date = datetime.strptime(week_to, "%Y/%m/%d")
                    if from_date > to_date:
                        messagebox.showwarning("入力エラー", "開始日は終了日より前の日付を指定してください。")
                        return
                except ValueError:
                    messagebox.showwarning("入力エラー", "日付は YYYY/MM/DD 形式で入力してください。\n例: 2025/01/06")
                    return
            self.show_step(3)

        elif self.current_step == 3:
            if not self.output_path.get():
                messagebox.showwarning("入力エラー", "出力先を選択してください。")
                return
            self.show_step(4)

        elif self.current_step == 4:
            self.execute()

    def execute(self):
        """集計を実行"""
        self.result = {
            "folder_path": self.folder_path.get(),
            "output_path": self.output_path.get(),
            "include_subfolders": self.include_subfolders.get(),
            "update_mode": self.update_mode.get(),
            "week_from": self.week_from.get().strip() or None,
            "week_to": self.week_to.get().strip() or None,
        }
        self.destroy()

    def cancel(self):
        """キャンセル"""
        self.result = None
        self.destroy()


def run_wizard():
    """ウィザードを実行して設定を取得"""
    app = WizardApp()
    app.mainloop()
    return app.result


# ===================================================================
#  ユーティリティ関数
# ===================================================================

def identify_team(filename):
    """ファイル名からチーム名を識別"""
    for pattern, team_name in TEAM_PATTERNS.items():
        if pattern in filename.upper():
            return team_name
    return "その他"


def _to_date(val):
    """セル値を日付文字列に変換（日付でなければNone）"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y/%m/%d")
    return None


def _to_date_obj(val):
    """セル値をdatetimeオブジェクトに変換"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        try:
            return datetime.strptime(val, "%Y/%m/%d")
        except ValueError:
            return None
    return None


def generate_date_range(start_date, end_date):
    """開始日から終了日までの日付リストを生成"""
    dates = []
    current = start_date
    while current <= end_date:
        dates.append(current)
        current += timedelta(days=1)
    return dates


def is_weekend(date_obj):
    """土日判定"""
    return date_obj.weekday() >= 5  # 5=土曜, 6=日曜


def load_cache(cache_file):
    """キャッシュファイルを読み込む"""
    if cache_file and os.path.exists(cache_file):
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_cache(cache_file, cache_data):
    """キャッシュファイルを保存する"""
    if cache_file:
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass


# ===================================================================
#  データ収集
# ===================================================================

def collect_data(folder_path, cache_file=None, include_subfolders=True):
    """フォルダ内の全Excelファイル → ITBシート → 全テストケースを収集"""

    records = []
    file_count = 0
    sheet_count = 0
    skipped_count = 0

    # キャッシュ読み込み
    cached_data = load_cache(cache_file)
    new_cache = {}

    # Excelファイルを収集
    excel_files = []
    if include_subfolders:
        for root, dirs, files in os.walk(folder_path):
            for filename in sorted(files):
                if filename.endswith(".xlsx") and not filename.startswith("~$"):
                    excel_files.append(os.path.join(root, filename))
    else:
        for filename in sorted(os.listdir(folder_path)):
            if filename.endswith(".xlsx") and not filename.startswith("~$"):
                excel_files.append(os.path.join(folder_path, filename))

    for filepath in excel_files:
        filename = os.path.basename(filepath)
        relative_path = os.path.relpath(filepath, folder_path)
        file_mtime = os.path.getmtime(filepath)

        # 差分チェック: ファイルの更新日時が変わっていなければキャッシュを使用
        if filepath in cached_data:
            cached_entry = cached_data[filepath]
            # 新形式: {mtime: ..., records: [...]} / 旧形式: float (mtime直接)
            if isinstance(cached_entry, dict) and cached_entry.get('mtime') == file_mtime:
                # キャッシュからレコードを復元
                cached_records = cached_entry.get('records', [])
                records.extend(cached_records)
                new_cache[filepath] = cached_entry
                skipped_count += 1
                print(f"  ⏭ {relative_path} (キャッシュ使用: {len(cached_records)}件)")
                continue
            # 旧形式のキャッシュは無視して再処理

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)

            # ITBで始まるシートがあるかチェック
            target_sheets = [s for s in wb.sheetnames if s.upper().startswith(SHEET_PREFIX.upper())]
            if not target_sheets:
                # 対象シートがない場合はスキップ（関係ないExcelファイル）
                wb.close()
                print(f"  ⏭ {relative_path} (対象シートなし - スキップ)")
                continue

            print(f"  📄 {relative_path}")
            file_count += 1

            # ファイル名からチーム名を識別
            team_name = identify_team(filename)
            file_records = []

            for ws_name in target_sheets:
                ws = wb[ws_name]
                sheet_count += 1
                case_count = 0

                for row in range(DATA_START_ROW, ws.max_row + 1):
                    test_id = ws.cell(row=row, column=COL_TEST_ID).value
                    if not test_id:
                        continue

                    jisshi_yotei   = ws.cell(row=row, column=COL_JISSHI_YOTEI).value
                    jisshi_jisseki = ws.cell(row=row, column=COL_JISSHI_JISSEKI).value
                    kensho_yotei   = ws.cell(row=row, column=COL_KENSHO_YOTEI).value
                    kensho_jisseki = ws.cell(row=row, column=COL_KENSHO_JISSEKI).value

                    record = {
                        "ファイル名": filepath,  # フルパスで記録
                        "シート名": ws_name,
                        "チーム名": team_name,
                        "テストID": str(test_id),
                        "実施者_予定": _to_date(jisshi_yotei),
                        "実施者_実績": _to_date(jisshi_jisseki),
                        "検証者_予定": _to_date(kensho_yotei),
                        "検証者_実績": _to_date(kensho_jisseki),
                    }
                    file_records.append(record)
                    case_count += 1

                print(f"     ✅ {ws_name} ({case_count}件) [チーム: {team_name}]")

            wb.close()

            # キャッシュに保存
            new_cache[filepath] = {
                'mtime': file_mtime,
                'records': file_records
            }
            records.extend(file_records)

        except Exception as e:
            print(f"     ⚠ エラー: {e}")

    # キャッシュを保存
    save_cache(cache_file, new_cache)

    print(f"\n  処理完了: {file_count}ファイル処理, {skipped_count}ファイルスキップ, {sheet_count}シート, {len(records)}レコード")
    return records


# ===================================================================
#  Excel出力
# ===================================================================

def write_excel(records, output_path, holidays=None, week_from=None, week_to=None):
    """ダッシュボード＋明細シート＋進捗サマリー（チーム別）＋祝日マスタをExcelに出力

    Args:
        records: テストケースレコードのリスト
        output_path: 出力ファイルパス
        holidays: 祝日リスト（省略時はデフォルト）
        week_from: 週集計の開始日（YYYY/MM/DD形式、省略可）
        week_to: 週集計の終了日（YYYY/MM/DD形式、省略可）
    """

    if holidays is None:
        holidays = DEFAULT_HOLIDAYS

    wb = openpyxl.Workbook()

    # --- 祝日マスタシート ---
    ws_holiday = wb.active
    ws_holiday.title = "祝日マスタ"
    _write_holiday_sheet(ws_holiday, holidays)

    # --- 明細シート ---
    ws_detail = wb.create_sheet("明細")
    detail_data_start_row = _write_detail_sheet(ws_detail, records)

    # --- チーム別にレコードを分類 ---
    team_records = defaultdict(list)
    for rec in records:
        team_records[rec["チーム名"]].append(rec)

    # チームリスト（ALL + 実際のチーム）
    teams_in_data = sorted(team_records.keys())

    # --- 遅延一覧シート ---
    ws_delayed = wb.create_sheet("要対応一覧")
    _write_delayed_sheet(ws_delayed, records, detail_data_start_row, len(records))

    # --- 進捗サマリーシート（ALL）- ダッシュボードより先に作成 ---
    summary_info = {}
    ws_summary_all = wb.create_sheet("進捗サマリー_ALL")
    summary_info["ALL"] = _write_summary_sheet(ws_summary_all, records, detail_data_start_row, len(records), holidays, "ALL")

    # --- 進捗サマリーシート（チーム別）- ダッシュボードより先に作成 ---
    for team_name in teams_in_data:
        team_recs = team_records[team_name]
        sheet_name = f"進捗サマリー_{team_name}"
        ws_team = wb.create_sheet(sheet_name)
        summary_info[team_name] = _write_summary_sheet(ws_team, team_recs, detail_data_start_row, len(records), holidays, team_name)

    # --- ダッシュボードシート（サマリーシート作成後に作成）---
    ws_dashboard = wb.create_sheet("ダッシュボード")
    _write_dashboard_sheet(ws_dashboard, summary_info, teams_in_data, wb, week_from, week_to)

    # シートの順序を調整
    # 目標の順序: ダッシュボード, 要対応一覧, 進捗サマリー_ALL, チーム別..., 明細, 祝日マスタ
    sheet_order = ["ダッシュボード", "要対応一覧", "進捗サマリー_ALL"]
    for team_name in teams_in_data:
        sheet_order.append(f"進捗サマリー_{team_name}")
    sheet_order.extend(["明細", "祝日マスタ"])

    # シートを並び替え
    for i, sheet_name in enumerate(sheet_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(sheet_name, offset=i - wb.sheetnames.index(sheet_name))

    # --- 保存 ---
    wb.save(output_path)
    print(f"\n  ✅ 出力完了: {output_path}")
    print(f"     ダッシュボード: 本日のサマリー")
    print(f"     明細シート: {len(records)}件")
    print(f"     サマリーシート: ALL + {len(teams_in_data)}チーム")


def _write_dashboard_sheet(ws, summary_info, team_list, wb, week_from=None, week_to=None):
    """ダッシュボードシート（5秒で状況把握）を作成

    構成:
        - 日次: 予定, 実績
        - 週次: 予定, 実績, 残数, 遅延（週範囲をセルに配置し数式で参照）
        - 総数: 総数, 予定, 実績, 残数, 遅延, 予定消化率, 実績消化率, 乖離, 状態

    Args:
        ws: ダッシュボードシートのワークシート
        summary_info: 各サマリーシートの参照情報
            {"ALL": {"total_row": N, "ref_row": M, ...}, "チーム名": {...}, ...}
        team_list: チーム名のリスト（表示順）
        wb: ワークブック（チャート参照用）
        week_from: 週集計の開始日（YYYY/MM/DD形式、省略可）
        week_to: 週集計の終了日（YYYY/MM/DD形式、省略可）
    """
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.axis import ChartLines
    from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
    from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    today = datetime.now()
    today_str = today.strftime("%Y/%m/%d")

    # グリッド線を非表示
    ws.sheet_view.showGridLines = False

    # --- 配色定義（指示書準拠） ---
    # 実施系: 青
    IMPL_SECTION_BG = "2B5797"      # セクションヘッダー
    IMPL_HEADER_BG = "4472C4"       # テーブルヘッダー
    IMPL_PLAN_COLOR = "4472C4"      # チャート予定線（青）
    IMPL_ACTUAL_COLOR = "ED7D31"    # チャート実績線（オレンジ）
    # 検証系: 緑
    VERIFY_SECTION_BG = "2E7D32"    # セクションヘッダー
    VERIFY_HEADER_BG = "548235"     # テーブルヘッダー
    VERIFY_PLAN_COLOR = "70AD47"    # チャート予定線（緑）
    VERIFY_ACTUAL_COLOR = "ED7D31"  # チャート実績線（オレンジ）
    # 共通
    COMMON_HEADER_BG = "6D6D6D"     # 共通テーブルヘッダー

    # --- タイトルエリア ---
    ws.merge_cells('A1:R1')
    title_cell = ws['A1']
    title_cell.value = "テスト進捗ダッシュボード"
    title_cell.font = Font(name="游ゴシック", size=18, bold=True, color="1B3A5C")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    # 基準日と週範囲を2行目に配置
    # A2: 基準日ラベル
    ws['A2'] = "基準日:"
    ws['A2'].font = Font(name="游ゴシック", size=11, bold=True, color="505050")
    ws['A2'].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 22

    # B2:D2: 基準日の値（結合セル）- 数式参照用
    ws.merge_cells('B2:D2')
    ws['B2'] = datetime.strptime(today_str, "%Y/%m/%d").date()
    ws['B2'].font = Font(name="游ゴシック", size=11, bold=True, color="2B5797")
    ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B2'].border = THIN_BORDER
    ws['B2'].number_format = "YYYY/MM/DD (AAA)"
    # 結合セル内の罫線
    ws['C2'].border = THIN_BORDER
    ws['D2'].border = THIN_BORDER

    # 週範囲セル
    # 週範囲ラベル
    ws['F2'] = "週範囲From:"
    ws['F2'].font = Font(name="游ゴシック", size=10, color="505050")
    ws['F2'].alignment = Alignment(horizontal="right", vertical="center")

    # 週From値セル（G2）- 数式参照用
    ws['G2'] = week_from if week_from else ""
    ws['G2'].font = Font(name="游ゴシック", size=9, bold=True, color="2B5797")
    ws['G2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['G2'].border = THIN_BORDER
    if week_from:
        ws['G2'].number_format = "YYYY/MM/DD"
    ws.column_dimensions['G'].width = 12

    ws['H2'] = "To:"
    ws['H2'].font = Font(name="游ゴシック", size=10, color="505050")
    ws['H2'].alignment = Alignment(horizontal="right", vertical="center")

    # 週To値セル（I2）- 数式参照用
    ws['I2'] = week_to if week_to else ""
    ws['I2'].font = Font(name="游ゴシック", size=9, bold=True, color="2B5797")
    ws['I2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['I2'].border = THIN_BORDER
    if week_to:
        ws['I2'].number_format = "YYYY/MM/DD"
    ws.column_dimensions['I'].width = 12

    # 週範囲セルの参照用アドレス（数式で使用）
    WEEK_FROM_CELL = "$G$2"
    WEEK_TO_CELL = "$I$2"

    # チーム順序（全体を先頭に）
    ordered_teams = ["全体"] + [t for t in team_list if t != "ALL"]

    def get_sheet_name(team):
        return "進捗サマリー_ALL" if team == "全体" else f"進捗サマリー_{team}"

    def get_info_key(team):
        return "ALL" if team == "全体" else team

    # 罫線スタイル（グループ境界用）
    GROUP_LEFT_BORDER = Border(
        left=MEDIUM_SOLID_SIDE,
        right=THIN_SOLID_SIDE,
        top=THIN_SOLID_SIDE,
        bottom=THIN_SOLID_SIDE,
    )
    GROUP_RIGHT_BORDER = Border(
        left=THIN_SOLID_SIDE,
        right=MEDIUM_SOLID_SIDE,
        top=THIN_SOLID_SIDE,
        bottom=THIN_SOLID_SIDE,
    )

    # データ行用の罫線設定関数（全セクション共通）
    def get_data_border(col):
        """列番号に応じたグループ境界罫線を返す"""
        if col == 1:  # チーム列
            return THIN_BORDER
        elif col == 2:  # 日次開始
            return Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
        elif col == 3:  # 日次終了
            return Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
        elif col == 4:  # 週次開始
            return Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
        elif col >= 5 and col <= 6:  # 週次中間
            return THIN_BORDER
        elif col == 7:  # 週次終了
            return Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
        elif col == 8:  # 総計開始
            return Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
        elif col >= 9 and col <= 15:  # 総計中間
            return THIN_BORDER
        elif col == 16:  # 総計終了
            return Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
        else:
            return THIN_BORDER

    # =================================================================
    # セクション1: 実施進捗
    # =================================================================
    row = 4
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = "■ 実施進捗"
    ws[f'A{row}'].font = Font(name="游ゴシック", size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color=IMPL_SECTION_BG, end_color=IMPL_SECTION_BG, fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
    ws[f'A{row}'].border = THIN_BORDER
    # マージされた全セルに罫線を適用
    for c in range(1, 17):
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.row_dimensions[row].height = 24

    # グループヘッダー行（日次 | 週次 | 総計）
    row += 1
    # チーム列
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=1).fill = PatternFill(start_color=COMMON_HEADER_BG, end_color=COMMON_HEADER_BG, fill_type="solid")
    ws.cell(row=row, column=1).border = THIN_BORDER

    # 日次グループ（B-C列）
    ws.merge_cells(f'B{row}:C{row}')
    ws.cell(row=row, column=2, value="日次")
    ws.cell(row=row, column=2).font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row, column=2).fill = PatternFill(start_color=IMPL_HEADER_BG, end_color=IMPL_HEADER_BG, fill_type="solid")
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=2).border = Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.cell(row=row, column=3).fill = PatternFill(start_color=IMPL_HEADER_BG, end_color=IMPL_HEADER_BG, fill_type="solid")
    ws.cell(row=row, column=3).border = Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)

    # 週次グループ（D-G列）- マージ前に全セルにスタイル設定
    impl_week_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    for c in range(4, 8):
        ws.cell(row=row, column=c).fill = impl_week_fill
        ws.cell(row=row, column=c).border = Border(left=THIN_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.merge_cells(f'D{row}:G{row}')
    ws.cell(row=row, column=4, value="週次")
    ws.cell(row=row, column=4).font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).border = Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.cell(row=row, column=7).border = Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)

    # 総計グループ（H-P列）- マージ前に全セルにスタイル設定
    impl_total_fill = PatternFill(start_color=IMPL_HEADER_BG, end_color=IMPL_HEADER_BG, fill_type="solid")
    for c in range(8, 17):
        ws.cell(row=row, column=c).fill = impl_total_fill
        ws.cell(row=row, column=c).border = Border(left=THIN_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.merge_cells(f'H{row}:P{row}')
    ws.cell(row=row, column=8, value="総計")
    ws.cell(row=row, column=8).font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=8).border = Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.cell(row=row, column=16).border = Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.row_dimensions[row].height = 20

    # サブヘッダー行（各項目名）
    row += 1

    # A5:A6を結合（グループヘッダー行とサブヘッダー行のチーム列）
    ws.merge_cells(f'A{row-1}:A{row}')
    ws.cell(row=row-1, column=1, value="チーム")
    ws.cell(row=row-1, column=1).font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row-1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    impl_headers = [
        "",  # チーム列は結合済みなのでスキップ
        # 日次（基準日ベース）
        "予定", "実績",
        # 週次（週範囲ベース）
        "予定", "実績", "残数", "遅延",
        # 総計
        "総数", "予定累計", "実績累計", "残数", "遅延",
        "予定消化率", "実績消化率", "乖離", "状態"
    ]
    for col, header in enumerate(impl_headers, 1):
        # col==1はA列結合済みのためスキップ
        if col == 1:
            continue
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name="游ゴシック", size=9, bold=True, color="FFFFFF")
        cell.alignment = HEADER_ALIGN

        # 列グループごとに色分けと罫線
        if col == 2:  # 日次開始
            cell.fill = PatternFill(start_color=IMPL_HEADER_BG, end_color=IMPL_HEADER_BG, fill_type="solid")
            cell.border = GROUP_LEFT_BORDER
        elif col == 3:  # 日次終了
            cell.fill = PatternFill(start_color=IMPL_HEADER_BG, end_color=IMPL_HEADER_BG, fill_type="solid")
            cell.border = GROUP_RIGHT_BORDER
        elif col == 4:  # 週次開始
            cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            cell.border = GROUP_LEFT_BORDER
        elif col == 7:  # 週次終了
            cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            cell.border = GROUP_RIGHT_BORDER
        elif col <= 6:  # 週次中間
            cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            cell.border = THIN_BORDER
        elif col == 8:  # 総計開始
            cell.fill = PatternFill(start_color=IMPL_HEADER_BG, end_color=IMPL_HEADER_BG, fill_type="solid")
            cell.border = GROUP_LEFT_BORDER
        elif col == 16:  # 総計終了
            cell.fill = PatternFill(start_color=IMPL_HEADER_BG, end_color=IMPL_HEADER_BG, fill_type="solid")
            cell.border = GROUP_RIGHT_BORDER
        else:  # 総計中間
            cell.fill = PatternFill(start_color=IMPL_HEADER_BG, end_color=IMPL_HEADER_BG, fill_type="solid")
            cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22

    impl_data_start = row + 1

    # 実施データ行
    for i, team in enumerate(ordered_teams):
        row += 1
        info_key = get_info_key(team)
        sheet_name = get_sheet_name(team)

        if info_key not in summary_info:
            continue

        info = summary_info[info_key]
        total_row = info["total_row"]
        ref_row = info["ref_row"]
        data_start_row = info["data_start_row"]
        data_end_row = info["data_end_row"]

        is_total_row = (i == 0)  # 全体行
        row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid") if is_total_row else None

        # A: チーム名
        cell = ws.cell(row=row, column=1, value=team)
        cell.font = Font(name="游ゴシック", size=10, bold=True)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        # B: 日次予定 = サマリーの基準日（$B$2）に該当する行のD列（実施予定）
        # INDEX/MATCHで$B$2の日付に対応する行を取得
        formula = f"=IFERROR(INDEX('{sheet_name}'!D{data_start_row}:D{data_end_row},MATCH($B$2,'{sheet_name}'!A{data_start_row}:A{data_end_row},0)),0)"
        cell = ws.cell(row=row, column=2, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(2)
        if row_fill:
            cell.fill = row_fill

        # C: 日次実績 = サマリーの基準日（$B$2）に該当する行のE列（実施実績）
        formula = f"=IFERROR(INDEX('{sheet_name}'!E{data_start_row}:E{data_end_row},MATCH($B$2,'{sheet_name}'!A{data_start_row}:A{data_end_row},0)),0)"
        cell = ws.cell(row=row, column=3, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(3)
        if row_fill:
            cell.fill = row_fill

        # D: 週予定 = SUMIFSで週範囲内の予定合計（常に数式を使用）
        formula = f"=IF(OR({WEEK_FROM_CELL}=\"\",{WEEK_TO_CELL}=\"\"),\"-\",SUMIFS('{sheet_name}'!D{data_start_row}:D{data_end_row},'{sheet_name}'!A{data_start_row}:A{data_end_row},\">=\"&{WEEK_FROM_CELL},'{sheet_name}'!A{data_start_row}:A{data_end_row},\"<=\"&{WEEK_TO_CELL}))"
        cell = ws.cell(row=row, column=4, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(4)
        if row_fill:
            cell.fill = row_fill

        # E: 週実績 = SUMIFSで週範囲内の実績合計（常に数式を使用）
        formula = f"=IF(OR({WEEK_FROM_CELL}=\"\",{WEEK_TO_CELL}=\"\"),\"-\",SUMIFS('{sheet_name}'!E{data_start_row}:E{data_end_row},'{sheet_name}'!A{data_start_row}:A{data_end_row},\">=\"&{WEEK_FROM_CELL},'{sheet_name}'!A{data_start_row}:A{data_end_row},\"<=\"&{WEEK_TO_CELL}))"
        cell = ws.cell(row=row, column=5, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(5)
        if row_fill:
            cell.fill = row_fill

        # F: 週残数 = 週予定 - 週実績（常に数式を使用）
        formula = f'=IF(OR({WEEK_FROM_CELL}="",{WEEK_TO_CELL}=""),"-",D{row}-E{row})'
        cell = ws.cell(row=row, column=6, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(6)
        if row_fill:
            cell.fill = row_fill

        # G: 週遅延 = 週残数（遅延がある場合のみ表示、常に数式を使用）
        formula = f'=IF(OR({WEEK_FROM_CELL}="",{WEEK_TO_CELL}=""),"-",IF(F{row}>0,F{row},0))'
        cell = ws.cell(row=row, column=7, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(7)
        if row_fill:
            cell.fill = row_fill

        # H: 総数 = サマリーの合計行D列
        cell = ws.cell(row=row, column=8, value=f"='{sheet_name}'!D{total_row}")
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(8)
        if row_fill:
            cell.fill = row_fill

        # I: 予定累計 = サマリーの基準日（$B$2）に該当する行のG列
        formula = f"=IFERROR(INDEX('{sheet_name}'!G{data_start_row}:G{data_end_row},MATCH($B$2,'{sheet_name}'!A{data_start_row}:A{data_end_row},0)),0)"
        cell = ws.cell(row=row, column=9, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(9)
        if row_fill:
            cell.fill = row_fill

        # J: 実績累計 = サマリーの基準日（$B$2）に該当する行のH列
        formula = f"=IFERROR(INDEX('{sheet_name}'!H{data_start_row}:H{data_end_row},MATCH($B$2,'{sheet_name}'!A{data_start_row}:A{data_end_row},0)),0)"
        cell = ws.cell(row=row, column=10, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(10)
        if row_fill:
            cell.fill = row_fill

        # K: 残数 = 総数 - 実績累計
        cell = ws.cell(row=row, column=11, value=f"=H{row}-J{row}")
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(11)
        if row_fill:
            cell.fill = row_fill

        # L: 遅延 = 予定累計 - 実績累計
        cell = ws.cell(row=row, column=12, value=f"=I{row}-J{row}")
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(12)
        if row_fill:
            cell.fill = row_fill

        # M: 予定消化率 = 予定累計 / 総数
        cell = ws.cell(row=row, column=13, value=f"=IF(H{row}=0,0,I{row}/H{row})")
        cell.number_format = "0.0%"
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(13)
        if row_fill:
            cell.fill = row_fill

        # N: 実績消化率 = 実績累計 / 総数
        cell = ws.cell(row=row, column=14, value=f"=IF(H{row}=0,0,J{row}/H{row})")
        cell.number_format = "0.0%"
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        # N: 実績消化率 = 実績累計 / 総数
        cell = ws.cell(row=row, column=14, value=f"=IF(H{row}=0,0,J{row}/H{row})")
        cell.number_format = "0.0%"
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(14)
        if row_fill:
            cell.fill = row_fill

        # O: 乖離 = 実績消化率 - 予定消化率
        cell = ws.cell(row=row, column=15, value=f"=N{row}-M{row}")
        cell.number_format = "0.0%"
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(15)
        if row_fill:
            cell.fill = row_fill

        # P: 状態
        cell = ws.cell(row=row, column=16, value=f'=IF(J{row}>=H{row},"完了",IF(L{row}>0,"遅延","順調"))')
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(16)
        if row_fill:
            cell.fill = row_fill

    impl_data_end = row

    # 実施の条件付き書式（遅延列と状態列）
    ws.conditional_formatting.add(
        f"L{impl_data_start}:L{impl_data_end}",
        FormulaRule(formula=[f'L{impl_data_start}>0'], fill=DANGER_FILL, font=DANGER_FONT)
    )
    ws.conditional_formatting.add(
        f"O{impl_data_start}:O{impl_data_end}",
        FormulaRule(formula=[f'O{impl_data_start}<0'], fill=DANGER_FILL, font=DANGER_FONT)
    )
    ws.conditional_formatting.add(
        f"P{impl_data_start}:P{impl_data_end}",
        FormulaRule(formula=[f'P{impl_data_start}="完了"'], fill=COMPLETE_FILL, font=COMPLETE_FONT)
    )
    ws.conditional_formatting.add(
        f"P{impl_data_start}:P{impl_data_end}",
        FormulaRule(formula=[f'P{impl_data_start}="遅延"'], fill=DANGER_FILL, font=DANGER_FONT)
    )
    ws.conditional_formatting.add(
        f"P{impl_data_start}:P{impl_data_end}",
        FormulaRule(formula=[f'P{impl_data_start}="順調"'], fill=OK_FILL, font=OK_FONT)
    )
    # 週遅延列の条件付き書式（常に適用）
    ws.conditional_formatting.add(
        f"G{impl_data_start}:G{impl_data_end}",
        FormulaRule(formula=[f'AND(ISNUMBER(G{impl_data_start}),G{impl_data_start}>0)'], fill=DANGER_FILL, font=DANGER_FONT)
    )

    # =================================================================
    # セクション2: 検証進捗
    # =================================================================
    row += 2
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = "■ 検証進捗"
    ws[f'A{row}'].font = Font(name="游ゴシック", size=12, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color=VERIFY_SECTION_BG, end_color=VERIFY_SECTION_BG, fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    # セクションヘッダー全体に罫線を適用
    for c in range(1, 17):
        ws.cell(row=row, column=c).border = THIN_BORDER

    # グループヘッダー行（日次 | 週次 | 総計）
    row += 1
    # チーム列
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=1).fill = PatternFill(start_color=COMMON_HEADER_BG, end_color=COMMON_HEADER_BG, fill_type="solid")
    ws.cell(row=row, column=1).border = THIN_BORDER

    # 日次グループ（B-C列）
    ws.merge_cells(f'B{row}:C{row}')
    verify_daily_fill = PatternFill(start_color=VERIFY_HEADER_BG, end_color=VERIFY_HEADER_BG, fill_type="solid")
    ws.cell(row=row, column=2, value="日次")
    ws.cell(row=row, column=2).font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row, column=2).fill = verify_daily_fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=2).border = Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.cell(row=row, column=3).fill = verify_daily_fill
    ws.cell(row=row, column=3).border = Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)

    # 週次グループ（D-G列）- マージ前に全セルにスタイル設定
    verify_week_fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
    for c in range(4, 8):
        ws.cell(row=row, column=c).fill = verify_week_fill
        ws.cell(row=row, column=c).border = Border(left=THIN_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.merge_cells(f'D{row}:G{row}')
    ws.cell(row=row, column=4, value="週次")
    ws.cell(row=row, column=4).font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=4).border = Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.cell(row=row, column=7).border = Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)

    # 総計グループ（H-P列）- マージ前に全セルにスタイル設定
    verify_total_fill = PatternFill(start_color=VERIFY_HEADER_BG, end_color=VERIFY_HEADER_BG, fill_type="solid")
    for c in range(8, 17):
        ws.cell(row=row, column=c).fill = verify_total_fill
        ws.cell(row=row, column=c).border = Border(left=THIN_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.merge_cells(f'H{row}:P{row}')
    ws.cell(row=row, column=8, value="総計")
    ws.cell(row=row, column=8).font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=8).border = Border(left=MEDIUM_SOLID_SIDE, right=THIN_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.cell(row=row, column=16).border = Border(left=THIN_SOLID_SIDE, right=MEDIUM_SOLID_SIDE, top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE)
    ws.row_dimensions[row].height = 20

    # サブヘッダー行（各項目名）
    row += 1

    # グループヘッダー行とサブヘッダー行のチーム列を結合
    ws.merge_cells(f'A{row-1}:A{row}')
    ws.cell(row=row-1, column=1, value="チーム")
    ws.cell(row=row-1, column=1).font = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    ws.cell(row=row-1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(impl_headers, 1):
        # col==1はA列結合済みのためスキップ
        if col == 1:
            continue
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name="游ゴシック", size=9, bold=True, color="FFFFFF")
        cell.alignment = HEADER_ALIGN

        # 列グループごとに色分けと罫線
        if col == 2:  # 日次開始
            cell.fill = PatternFill(start_color=VERIFY_HEADER_BG, end_color=VERIFY_HEADER_BG, fill_type="solid")
            cell.border = GROUP_LEFT_BORDER
        elif col == 3:  # 日次終了
            cell.fill = PatternFill(start_color=VERIFY_HEADER_BG, end_color=VERIFY_HEADER_BG, fill_type="solid")
            cell.border = GROUP_RIGHT_BORDER
        elif col == 4:  # 週次開始
            cell.fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
            cell.border = GROUP_LEFT_BORDER
        elif col == 7:  # 週次終了
            cell.fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
            cell.border = GROUP_RIGHT_BORDER
        elif col <= 6:  # 週次中間
            cell.fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
            cell.border = THIN_BORDER
        elif col == 8:  # 総計開始
            cell.fill = PatternFill(start_color=VERIFY_HEADER_BG, end_color=VERIFY_HEADER_BG, fill_type="solid")
            cell.border = GROUP_LEFT_BORDER
        elif col == 16:  # 総計終了
            cell.fill = PatternFill(start_color=VERIFY_HEADER_BG, end_color=VERIFY_HEADER_BG, fill_type="solid")
            cell.border = GROUP_RIGHT_BORDER
        else:  # 総計中間
            cell.fill = PatternFill(start_color=VERIFY_HEADER_BG, end_color=VERIFY_HEADER_BG, fill_type="solid")
            cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22

    verify_data_start = row + 1

    # 検証データ行
    for i, team in enumerate(ordered_teams):
        row += 1
        info_key = get_info_key(team)
        sheet_name = get_sheet_name(team)

        if info_key not in summary_info:
            continue

        info = summary_info[info_key]
        total_row = info["total_row"]
        ref_row = info["ref_row"]
        data_start_row = info["data_start_row"]
        data_end_row = info["data_end_row"]

        is_total_row = (i == 0)
        row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") if is_total_row else None

        # A: チーム名
        cell = ws.cell(row=row, column=1, value=team)
        cell.font = Font(name="游ゴシック", size=10, bold=True)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill

        # B: 日次予定 = サマリーの基準日（$B$2）に該当する行のK列（検証予定）
        formula = f"=IFERROR(INDEX('{sheet_name}'!K{data_start_row}:K{data_end_row},MATCH($B$2,'{sheet_name}'!A{data_start_row}:A{data_end_row},0)),0)"
        cell = ws.cell(row=row, column=2, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(2)
        if row_fill:
            cell.fill = row_fill

        # C: 日次実績 = サマリーの基準日（$B$2）に該当する行のL列（検証実績）
        formula = f"=IFERROR(INDEX('{sheet_name}'!L{data_start_row}:L{data_end_row},MATCH($B$2,'{sheet_name}'!A{data_start_row}:A{data_end_row},0)),0)"
        cell = ws.cell(row=row, column=3, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(3)
        if row_fill:
            cell.fill = row_fill

        # D: 週予定 = SUMIFSで週範囲内の予定合計（検証K列、常に数式を使用）
        formula = f"=IF(OR({WEEK_FROM_CELL}=\"\",{WEEK_TO_CELL}=\"\"),\"-\",SUMIFS('{sheet_name}'!K{data_start_row}:K{data_end_row},'{sheet_name}'!A{data_start_row}:A{data_end_row},\">=\"&{WEEK_FROM_CELL},'{sheet_name}'!A{data_start_row}:A{data_end_row},\"<=\"&{WEEK_TO_CELL}))"
        cell = ws.cell(row=row, column=4, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(4)
        if row_fill:
            cell.fill = row_fill

        # E: 週実績 = SUMIFSで週範囲内の実績合計（検証L列、常に数式を使用）
        formula = f"=IF(OR({WEEK_FROM_CELL}=\"\",{WEEK_TO_CELL}=\"\"),\"-\",SUMIFS('{sheet_name}'!L{data_start_row}:L{data_end_row},'{sheet_name}'!A{data_start_row}:A{data_end_row},\">=\"&{WEEK_FROM_CELL},'{sheet_name}'!A{data_start_row}:A{data_end_row},\"<=\"&{WEEK_TO_CELL}))"
        cell = ws.cell(row=row, column=5, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(5)
        if row_fill:
            cell.fill = row_fill

        # F: 週残数 = 週予定 - 週実績（常に数式を使用）
        formula = f'=IF(OR({WEEK_FROM_CELL}="",{WEEK_TO_CELL}=""),"-",D{row}-E{row})'
        cell = ws.cell(row=row, column=6, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(6)
        if row_fill:
            cell.fill = row_fill

        # G: 週遅延 = 週残数（遅延がある場合のみ表示、常に数式を使用）
        formula = f'=IF(OR({WEEK_FROM_CELL}="",{WEEK_TO_CELL}=""),"-",IF(F{row}>0,F{row},0))'
        cell = ws.cell(row=row, column=7, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(7)
        if row_fill:
            cell.fill = row_fill

        # H: 総数 = サマリーの合計行K列（検証予定の合計）
        cell = ws.cell(row=row, column=8, value=f"='{sheet_name}'!K{total_row}")
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(8)
        if row_fill:
            cell.fill = row_fill

        # I: 予定累計 = サマリーの基準日（$B$2）に該当する行のN列
        formula = f"=IFERROR(INDEX('{sheet_name}'!N{data_start_row}:N{data_end_row},MATCH($B$2,'{sheet_name}'!A{data_start_row}:A{data_end_row},0)),0)"
        cell = ws.cell(row=row, column=9, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(9)
        if row_fill:
            cell.fill = row_fill

        # J: 実績累計 = サマリーの基準日（$B$2）に該当する行のO列
        formula = f"=IFERROR(INDEX('{sheet_name}'!O{data_start_row}:O{data_end_row},MATCH($B$2,'{sheet_name}'!A{data_start_row}:A{data_end_row},0)),0)"
        cell = ws.cell(row=row, column=10, value=formula)
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(10)
        if row_fill:
            cell.fill = row_fill

        # K: 残数 = 総数 - 実績累計
        cell = ws.cell(row=row, column=11, value=f"=H{row}-J{row}")
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(11)
        if row_fill:
            cell.fill = row_fill

        # L: 遅延 = 予定累計 - 実績累計
        cell = ws.cell(row=row, column=12, value=f"=I{row}-J{row}")
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(12)
        if row_fill:
            cell.fill = row_fill

        # M: 予定消化率 = 予定累計 / 総数
        cell = ws.cell(row=row, column=13, value=f"=IF(H{row}=0,0,I{row}/H{row})")
        cell.number_format = "0.0%"
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(13)
        if row_fill:
            cell.fill = row_fill

        # N: 実績消化率 = 実績累計 / 総数
        cell = ws.cell(row=row, column=14, value=f"=IF(H{row}=0,0,J{row}/H{row})")
        cell.number_format = "0.0%"
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(14)
        if row_fill:
            cell.fill = row_fill

        # O: 乖離 = 実績消化率 - 予定消化率
        cell = ws.cell(row=row, column=15, value=f"=N{row}-M{row}")
        cell.number_format = "0.0%"
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(15)
        if row_fill:
            cell.fill = row_fill

        # P: 状態
        cell = ws.cell(row=row, column=16, value=f'=IF(J{row}>=H{row},"完了",IF(L{row}>0,"遅延","順調"))')
        cell.alignment = DATA_ALIGN_CENTER
        cell.border = get_data_border(16)
        if row_fill:
            cell.fill = row_fill

    verify_data_end = row

    # 検証の条件付き書式
    ws.conditional_formatting.add(
        f"L{verify_data_start}:L{verify_data_end}",
        FormulaRule(formula=[f'L{verify_data_start}>0'], fill=DANGER_FILL, font=DANGER_FONT)
    )
    ws.conditional_formatting.add(
        f"O{verify_data_start}:O{verify_data_end}",
        FormulaRule(formula=[f'O{verify_data_start}<0'], fill=DANGER_FILL, font=DANGER_FONT)
    )
    ws.conditional_formatting.add(
        f"P{verify_data_start}:P{verify_data_end}",
        FormulaRule(formula=[f'P{verify_data_start}="完了"'], fill=COMPLETE_FILL, font=COMPLETE_FONT)
    )
    ws.conditional_formatting.add(
        f"P{verify_data_start}:P{verify_data_end}",
        FormulaRule(formula=[f'P{verify_data_start}="遅延"'], fill=DANGER_FILL, font=DANGER_FONT)
    )
    ws.conditional_formatting.add(
        f"P{verify_data_start}:P{verify_data_end}",
        FormulaRule(formula=[f'P{verify_data_start}="順調"'], fill=OK_FILL, font=OK_FONT)
    )
    # 週遅延列の条件付き書式（常に適用）
    ws.conditional_formatting.add(
        f"G{verify_data_start}:G{verify_data_end}",
        FormulaRule(formula=[f'AND(ISNUMBER(G{verify_data_start}),G{verify_data_start}>0)'], fill=DANGER_FILL, font=DANGER_FONT)
    )

    # =================================================================
    # セクション3: 進捗推移チャート
    # =================================================================
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont

    row += 2
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = "■ 進捗推移チャート（左: 実施 / 右: 検証）　※実施予定=青、検証予定=緑、実績=オレンジ"
    ws[f'A{row}'].font = Font(name="游ゴシック", size=12, bold=True, color="505050")
    ws[f'A{row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24

    # チーム順序（指示書準拠: 全体→オンライン→バッチ→基盤→運用→その他）
    chart_team_order = ["全体", "オンライン", "バッチ", "基盤", "運用", "その他"]

    # pt → EMU変換（1pt = 12700 EMU）
    PT_TO_EMU = 12700

    # チャートサイズ（pt単位、cm単位両方で保持）
    # 1pt = 0.0352778cm
    PT_TO_CM = 0.0352778

    # 大外の枠サイズ（pt）- 高さを大きくして余白を確保
    IMPL_CHART_WIDTH_PT = 380    # 実施チャート幅
    IMPL_CHART_HEIGHT_PT = 320   # 実施チャート高さ
    VERIFY_CHART_WIDTH_PT = 450  # 検証チャート幅
    VERIFY_CHART_HEIGHT_PT = 320 # 検証チャート高さ

    # cm単位に変換（chart.width/heightに使用）
    IMPL_CHART_WIDTH = IMPL_CHART_WIDTH_PT * PT_TO_CM
    IMPL_CHART_HEIGHT = IMPL_CHART_HEIGHT_PT * PT_TO_CM
    VERIFY_CHART_WIDTH = VERIFY_CHART_WIDTH_PT * PT_TO_CM
    VERIFY_CHART_HEIGHT = VERIFY_CHART_HEIGHT_PT * PT_TO_CM

    # チャート配置（pt）
    IMPL_CHART_LEFT_PT = 0       # 実施チャート左端
    VERIFY_CHART_LEFT_PT = 388   # 検証チャート左端（380 + 8pt gap）
    CHART_TOP_START_PT = 520     # 最初のチャート（全体）のtop位置（新レイアウトに合わせて調整）
    CHART_VERTICAL_GAP_PT = 328  # チャート縦間隔（320 + 8pt gap）

    # PlotAreaレイアウト比率
    PLOT_IMPL = {"x": 0.02, "y": 0.15, "w": 0.96, "h": 0.60}
    PLOT_VER = {"x": 0.02, "y": 0.15, "w": 0.96, "h": 0.60}

    # フォント色
    FONT_COLOR = "595959"

    chart_count = 0
    for team in chart_team_order:
        info_key = "ALL" if team == "全体" else team
        sheet_name = "進捗サマリー_ALL" if team == "全体" else f"進捗サマリー_{team}"

        if info_key not in summary_info:
            continue

        info = summary_info[info_key]
        # データ範囲（合計行を除外するため、data_start_row=6から開始）
        data_start = info["data_start_row"]
        data_end = info["data_end_row"]

        # サマリーシートを取得
        if sheet_name not in wb.sheetnames:
            continue
        summary_ws = wb[sheet_name]

        # チャート共通フォント設定
        def setup_chart_style(chart, chart_type):
            """チャートのフォント・目盛り線・レイアウト設定を行う"""
            from openpyxl.chart.shapes import GraphicalProperties
            from openpyxl.drawing.line import LineProperties
            from openpyxl.chart.layout import Layout, ManualLayout

            # タイトルフォント（10pt、太字）
            title_font = DrawingFont(typeface="ＭＳ Ｐゴシック")
            title_cp = CharacterProperties(latin=title_font, sz=1000, b=True)
            title_cp.solidFill = FONT_COLOR
            title_paragraph = Paragraph(pPr=ParagraphProperties(defRPr=title_cp), endParaRPr=title_cp)
            chart.title.txPr = RichText(p=[title_paragraph])

            # 凡例フォント（8pt）
            legend_font = DrawingFont(typeface="ＭＳ Ｐゴシック")
            legend_cp = CharacterProperties(latin=legend_font, sz=800)
            legend_cp.solidFill = FONT_COLOR
            legend_paragraph = Paragraph(pPr=ParagraphProperties(defRPr=legend_cp), endParaRPr=legend_cp)
            chart.legend.txPr = RichText(p=[legend_paragraph])

            # X軸フォント（7pt）
            x_font = DrawingFont(typeface="ＭＳ Ｐゴシック")
            x_cp = CharacterProperties(latin=x_font, sz=700)
            x_cp.solidFill = FONT_COLOR
            x_paragraph = Paragraph(pPr=ParagraphProperties(defRPr=x_cp), endParaRPr=x_cp)
            chart.x_axis.txPr = RichText(p=[x_paragraph])

            # Y軸フォント（8pt）
            y_font = DrawingFont(typeface="ＭＳ Ｐゴシック")
            y_cp = CharacterProperties(latin=y_font, sz=800)
            y_cp.solidFill = FONT_COLOR
            y_paragraph = Paragraph(pPr=ParagraphProperties(defRPr=y_cp), endParaRPr=y_cp)
            chart.y_axis.txPr = RichText(p=[y_paragraph])

            # 目盛り線を薄いグレーに設定
            chart.y_axis.majorGridlines = ChartLines()
            chart.y_axis.majorGridlines.spPr = GraphicalProperties(
                ln=LineProperties(solidFill="D0D0D0", w=9525)
            )

            # X軸の主目盛り線（縦線）- 通常は非表示だが念のため
            chart.x_axis.majorGridlines = None

            # チャートの角を直角にする（角丸をなくす）
            chart.roundedCorners = False

            # PlotAreaのレイアウトを明示指定（chart_typeで分岐）
            # これが余白確保の核心
            if chart_type == "実施":
                plot_layout = PLOT_IMPL
            else:  # 検証
                plot_layout = PLOT_VER

            # ★重要: chart.layoutを設定する（chart.plot_area.layoutではない）
            # openpyxlは保存時にchart.layoutをplot_area.layoutにコピーする
            chart.layout = Layout(
                manualLayout=ManualLayout(
                    layoutTarget="inner", # PlotAreaの内側を対象
                    xMode="edge",         # 位置モード: edge（絶対位置）
                    yMode="edge",         # 位置モード: edge（絶対位置）
                    wMode="factor",       # サイズモード: factor（比率）
                    hMode="factor",       # サイズモード: factor（比率）
                    x=plot_layout["x"],   # 左からの位置（比率）
                    y=plot_layout["y"],   # 上からの位置（比率）= 33/250 = 0.132
                    w=plot_layout["w"],   # 幅（比率）
                    h=plot_layout["h"],   # 高さ（比率）= 189.5/250 = 0.758
                )
            )

        # 実施チャート（左）
        impl_chart = LineChart()
        impl_chart.title = f"{team} - 実施"
        impl_chart.style = 10
        impl_chart.y_axis.delete = False  # Y軸を表示
        impl_chart.x_axis.delete = False  # X軸を表示
        impl_chart.width = IMPL_CHART_WIDTH    # 380pt → cm
        impl_chart.height = IMPL_CHART_HEIGHT  # 250pt → cm
        impl_chart.legend.position = 'b'

        # データ参照（G列: 予定累計、H列: 実績累計、A列: 日付）
        impl_plan_data = Reference(summary_ws, min_col=7, min_row=data_start, max_row=data_end)
        impl_actual_data = Reference(summary_ws, min_col=8, min_row=data_start, max_row=data_end)
        impl_dates = Reference(summary_ws, min_col=1, min_row=data_start, max_row=data_end)

        impl_chart.add_data(impl_plan_data)
        impl_chart.add_data(impl_actual_data)
        impl_chart.set_categories(impl_dates)

        # 系列名と色を設定（実施予定=青・破線、実績=オレンジ・実線）
        if len(impl_chart.series) >= 1:
            impl_chart.series[0].tx = SeriesLabel(v="予定")
            impl_chart.series[0].graphicalProperties.line.solidFill = IMPL_PLAN_COLOR
            impl_chart.series[0].graphicalProperties.line.width = 25400  # 2pt = 25400 EMUs
            impl_chart.series[0].graphicalProperties.line.dashStyle = "sysDash"  # 短い間隔の破線
            impl_chart.series[0].marker.symbol = "none"  # マーカーなし
            impl_chart.series[0].smooth = False  # 直線で結ぶ（曲線にしない）
        if len(impl_chart.series) >= 2:
            impl_chart.series[1].tx = SeriesLabel(v="実績")
            impl_chart.series[1].graphicalProperties.line.solidFill = IMPL_ACTUAL_COLOR
            impl_chart.series[1].graphicalProperties.line.width = 25400  # 2pt
            impl_chart.series[1].marker.symbol = "none"  # マーカーなし
            impl_chart.series[1].smooth = False  # 直線で結ぶ（曲線にしない）

        # スタイル設定を適用
        setup_chart_style(impl_chart, "実施")

        # 実施チャートの配置（AbsoluteAnchorでpt単位制御）
        # top = 460pt + chart_count * 258pt
        impl_top_pt = CHART_TOP_START_PT + chart_count * CHART_VERTICAL_GAP_PT
        impl_chart.anchor = AbsoluteAnchor(
            pos=XDRPoint2D(
                x=IMPL_CHART_LEFT_PT * PT_TO_EMU,  # 0pt
                y=impl_top_pt * PT_TO_EMU
            ),
            ext=XDRPositiveSize2D(
                cx=IMPL_CHART_WIDTH_PT * PT_TO_EMU,  # 380pt
                cy=IMPL_CHART_HEIGHT_PT * PT_TO_EMU  # 250pt
            )
        )
        ws.add_chart(impl_chart)

        # 検証チャート（右）
        verify_chart = LineChart()
        verify_chart.title = f"{team} - 検証"
        verify_chart.style = 10
        verify_chart.y_axis.delete = False  # Y軸を表示
        verify_chart.x_axis.delete = False  # X軸を表示
        verify_chart.width = VERIFY_CHART_WIDTH    # 450pt → cm
        verify_chart.height = VERIFY_CHART_HEIGHT  # 250pt → cm
        verify_chart.legend.position = 'b'

        # データ参照（N列: 予定累計、O列: 実績累計、A列: 日付）
        verify_plan_data = Reference(summary_ws, min_col=14, min_row=data_start, max_row=data_end)
        verify_actual_data = Reference(summary_ws, min_col=15, min_row=data_start, max_row=data_end)
        verify_dates = Reference(summary_ws, min_col=1, min_row=data_start, max_row=data_end)

        verify_chart.add_data(verify_plan_data)
        verify_chart.add_data(verify_actual_data)
        verify_chart.set_categories(verify_dates)

        # 系列名と色を設定（検証予定=緑、実績=オレンジ）
        # 系列名と色を設定（検証予定=緑・破線、実績=オレンジ・実線）
        if len(verify_chart.series) >= 1:
            verify_chart.series[0].tx = SeriesLabel(v="予定")
            verify_chart.series[0].graphicalProperties.line.solidFill = VERIFY_PLAN_COLOR
            verify_chart.series[0].graphicalProperties.line.width = 25400  # 2pt
            verify_chart.series[0].graphicalProperties.line.dashStyle = "sysDash"  # 短い間隔の破線
            verify_chart.series[0].marker.symbol = "none"  # マーカーなし
            verify_chart.series[0].smooth = False  # 直線で結ぶ（曲線にしない）
        if len(verify_chart.series) >= 2:
            verify_chart.series[1].tx = SeriesLabel(v="実績")
            verify_chart.series[1].graphicalProperties.line.solidFill = VERIFY_ACTUAL_COLOR
            verify_chart.series[1].graphicalProperties.line.width = 25400  # 2pt
            verify_chart.series[1].marker.symbol = "none"  # マーカーなし
            verify_chart.series[1].smooth = False  # 直線で結ぶ（曲線にしない）

        # スタイル設定を適用
        setup_chart_style(verify_chart, "検証")

        # 検証チャートの配置（AbsoluteAnchorでpt単位制御）
        # left = 388pt（実施380pt + 8pt gap）
        verify_top_pt = CHART_TOP_START_PT + chart_count * CHART_VERTICAL_GAP_PT
        verify_chart.anchor = AbsoluteAnchor(
            pos=XDRPoint2D(
                x=VERIFY_CHART_LEFT_PT * PT_TO_EMU,  # 388pt
                y=verify_top_pt * PT_TO_EMU
            ),
            ext=XDRPositiveSize2D(
                cx=VERIFY_CHART_WIDTH_PT * PT_TO_EMU,  # 450pt
                cy=VERIFY_CHART_HEIGHT_PT * PT_TO_EMU  # 250pt
            )
        )
        ws.add_chart(verify_chart)

        chart_count += 1

    # --- 列幅設定（新しい16列構成に対応）---
    ws.column_dimensions['A'].width = 10   # チーム
    ws.column_dimensions['B'].width = 8    # 日次予定
    ws.column_dimensions['C'].width = 8    # 日次実績
    ws.column_dimensions['D'].width = 8    # 週予定
    ws.column_dimensions['E'].width = 8    # 週実績
    ws.column_dimensions['F'].width = 8    # 週残数
    ws.column_dimensions['G'].width = 8    # 週遅延
    ws.column_dimensions['H'].width = 8    # 総数
    ws.column_dimensions['I'].width = 10   # 予定累計
    ws.column_dimensions['J'].width = 10   # 実績累計
    ws.column_dimensions['K'].width = 8    # 残数
    ws.column_dimensions['L'].width = 8    # 遅延
    ws.column_dimensions['M'].width = 10   # 予定消化率
    ws.column_dimensions['N'].width = 10   # 実績消化率
    ws.column_dimensions['O'].width = 8    # 乖離
    ws.column_dimensions['P'].width = 8    # 状態

    # 印刷設定
    ws.print_title_rows = '1:2'


def _write_delayed_sheet(ws, records, detail_start_row, total_records):
    """要対応一覧シート（遅延レコードの抽出）を作成"""

    # グリッド線を非表示
    ws.sheet_view.showGridLines = False

    detail_last_row = detail_start_row + total_records - 1
    today = datetime.now()
    today_str = today.strftime("%Y/%m/%d")

    # 遅延レコードを抽出
    delayed_records = []
    for rec in records:
        jisshi_yotei = _to_date_obj(rec["実施者_予定"])
        jisshi_jisseki = rec["実施者_実績"]
        kensho_yotei = _to_date_obj(rec["検証者_予定"])
        kensho_jisseki = rec["検証者_実績"]

        # 実施遅延: 予定日<=今日 かつ 実績なし
        jisshi_delayed = jisshi_yotei and jisshi_yotei <= today and not jisshi_jisseki
        # 検証遅延: 予定日<=今日 かつ 実績なし
        kensho_delayed = kensho_yotei and kensho_yotei <= today and not kensho_jisseki

        if jisshi_delayed or kensho_delayed:
            rec_copy = rec.copy()
            rec_copy['実施遅延'] = jisshi_delayed
            rec_copy['検証遅延'] = kensho_delayed
            delayed_records.append(rec_copy)

    # タイトル
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"要対応一覧（{len(delayed_records)}件）"
    title_cell.font = Font(name="游ゴシック", size=14, bold=True, color="8B0000")
    title_cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # 基準日
    ws['A2'] = f"基準日: {today_str}"
    ws['A2'].font = Font(name="游ゴシック", size=11, color="666666")

    # サマリー
    row = 4
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "■ 遅延サマリー"
    ws[f'A{row}'].font = DASHBOARD_SECTION_FONT
    ws[f'A{row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    row += 1
    summary_headers = ["項目", "遅延件数"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    row += 1
    ws.cell(row=row, column=1, value="実施者_遅延").border = THIN_BORDER
    ws.cell(row=row, column=1).alignment = DATA_ALIGN_CENTER
    ws.cell(row=row, column=2, value=f'=COUNTIF(明細!$H${detail_start_row}:$H${detail_last_row},"遅延")').border = THIN_BORDER
    ws.cell(row=row, column=2).alignment = DATA_ALIGN_CENTER
    ws.conditional_formatting.add(f"B{row}", FormulaRule(formula=[f'B{row}>0'], fill=DANGER_FILL, font=DANGER_FONT))

    row += 1
    ws.cell(row=row, column=1, value="検証者_遅延").border = THIN_BORDER
    ws.cell(row=row, column=1).alignment = DATA_ALIGN_CENTER
    ws.cell(row=row, column=2, value=f'=COUNTIF(明細!$K${detail_start_row}:$K${detail_last_row},"遅延")').border = THIN_BORDER
    ws.cell(row=row, column=2).alignment = DATA_ALIGN_CENTER
    ws.conditional_formatting.add(f"B{row}", FormulaRule(formula=[f'B{row}>0'], fill=DANGER_FILL, font=DANGER_FONT))

    # 遅延一覧
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "■ 遅延一覧"
    ws[f'A{row}'].font = DASHBOARD_SECTION_FONT
    ws[f'A{row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    row += 1
    list_headers = ["No.", "チーム名", "シート名", "テストID", "実施予定", "実施実績", "検証予定", "検証実績"]
    for col, header in enumerate(list_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    data_start = row + 1
    for i, rec in enumerate(delayed_records):
        row += 1
        values = [
            i + 1,
            rec["チーム名"],
            rec["シート名"],
            rec["テストID"],
            rec["実施者_予定"] or "－",
            rec["実施者_実績"] or "未完了",
            rec["検証者_予定"] or "－",
            rec["検証者_実績"] or "未完了",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN_CENTER

            # 遅延中のセルを強調
            if col == 6 and rec['実施遅延']:  # 実施実績
                cell.fill = DANGER_FILL
                cell.font = DANGER_FONT
            elif col == 8 and rec['検証遅延']:  # 検証実績
                cell.fill = DANGER_FILL
                cell.font = DANGER_FONT

    # テーブル作成
    if delayed_records:
        data_end = data_start + len(delayed_records) - 1
        table_ref = f"A{data_start - 1}:H{data_end}"
        try:
            table = Table(displayName="遅延一覧テーブル", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium3",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        except Exception:
            pass  # テーブル作成に失敗しても続行

    # 遅延がない場合
    if not delayed_records:
        row += 1
        ws.cell(row=row, column=1, value="遅延しているテストケースはありません").font = Font(name="游ゴシック", size=11, color="2E7D32", italic=True)

    # 列幅設定
    delayed_widths = [8, 14, 20, 18, 14, 14, 14, 14]
    for i, w in enumerate(delayed_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_holiday_sheet(ws, holidays):
    """祝日マスタシートを作成

    Args:
        holidays: 辞書形式 {日付: 祝日名} または リスト形式 [日付, ...]
    """

    # グリッド線を非表示
    ws.sheet_view.showGridLines = False

    # タイトル
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = "祝日マスタ"
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # 説明
    ws['A2'] = "※ 祝日を追加・編集してください。日付形式: YYYY/MM/DD"
    ws['A2'].font = Font(name="游ゴシック", size=9, color="666666")

    # ヘッダー
    headers = ["日付", "祝日名", "備考"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 祝日データ（辞書形式またはリスト形式に対応）
    if isinstance(holidays, dict):
        # 辞書形式: {日付: 祝日名}
        sorted_holidays = sorted(holidays.items(), key=lambda x: x[0])
        for i, (date_str, name) in enumerate(sorted_holidays):
            row = i + 5
            ws.cell(row=row, column=1, value=date_str).border = THIN_BORDER
            ws.cell(row=row, column=2, value=name).border = THIN_BORDER
            ws.cell(row=row, column=3, value="").border = THIN_BORDER
    else:
        # リスト形式: [日付, ...]（後方互換性）
        for i, holiday in enumerate(holidays):
            row = i + 5
            ws.cell(row=row, column=1, value=holiday).border = THIN_BORDER
            ws.cell(row=row, column=2, value="").border = THIN_BORDER
            ws.cell(row=row, column=3, value="").border = THIN_BORDER

    # 列幅
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25

    # ヘッダー固定（行4まで固定）
    ws.freeze_panes = 'A5'


def _write_detail_sheet(ws, records):
    """明細シートを作成（テーブル形式）"""

    # グリッド線を非表示
    ws.sheet_view.showGridLines = False

    # タイトル (A1)
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = "テスト進捗明細"
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # 基準日ラベルと値 (K2, L2) - ダッシュボードのB2を参照
    ws['K2'] = "基準日:"
    ws['K2'].font = Font(name="游ゴシック", size=11, bold=True)
    ws['K2'].alignment = DATA_ALIGN_RIGHT

    ws['L2'] = "=ダッシュボード!$B$2"
    ws['L2'].font = REF_DATE_FONT
    ws['L2'].fill = REF_DATE_FILL
    ws['L2'].alignment = DATA_ALIGN_CENTER
    ws['L2'].border = THIN_BORDER
    ws['L2'].number_format = "YYYY/MM/DD"

    # ヘッダー行 (row 4)
    header_row = 4
    detail_headers = [
        "No.", "ファイル名", "シート名", "チーム名", "テストID",
        "実施者_予定", "実施者_実績", "実施者_状況",
        "検証者_予定", "検証者_実績", "検証者_状況",
        "進捗状況",
    ]

    for col, header in enumerate(detail_headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # データ行
    data_start_row = header_row + 1
    for i, rec in enumerate(records):
        row = data_start_row + i

        # 実施者状況（実績があれば完了、なければ予定日と基準日を比較）
        jisshi_status = '=IF(G{row}<>"","完了",IF(F{row}="","－",IF(F{row}<=$L$2,"遅延","予定")))'.format(row=row)
        # 検証者状況
        kensho_status = '=IF(J{row}<>"","完了",IF(I{row}="","－",IF(I{row}<=$L$2,"遅延","予定")))'.format(row=row)
        # 全体進捗
        overall_status = '=IF(AND(H{row}="完了",K{row}="完了"),"完了",IF(OR(H{row}="遅延",K{row}="遅延"),"遅延","進行中"))'.format(row=row)

        values = [
            i + 1,
            rec["ファイル名"],
            rec["シート名"],
            rec["チーム名"],
            rec["テストID"],
            rec["実施者_予定"] if rec["実施者_予定"] else "",
            rec["実施者_実績"] if rec["実施者_実績"] else "",
            jisshi_status,
            rec["検証者_予定"] if rec["検証者_予定"] else "",
            rec["検証者_実績"] if rec["検証者_実績"] else "",
            kensho_status,
            overall_status,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if col == 1:
                cell.alignment = DATA_ALIGN_CENTER
            elif col in (4, 5, 6, 7, 8, 9, 10, 11, 12):
                cell.alignment = DATA_ALIGN_CENTER
            else:
                cell.alignment = DATA_ALIGN_LEFT

    # テーブル作成
    if records:
        table_ref = f"A{header_row}:L{data_start_row + len(records) - 1}"
        table = Table(displayName="明細テーブル", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # 条件付き書式: 状況列（H, K, L列）- 4色システム
    last_row = data_start_row + len(records) - 1 if records else data_start_row

    # 完了 = 緑
    ws.conditional_formatting.add(
        f"H{data_start_row}:H{last_row}",
        FormulaRule(formula=[f'H{data_start_row}="完了"'], fill=COMPLETE_FILL, font=COMPLETE_FONT)
    )
    ws.conditional_formatting.add(
        f"K{data_start_row}:K{last_row}",
        FormulaRule(formula=[f'K{data_start_row}="完了"'], fill=COMPLETE_FILL, font=COMPLETE_FONT)
    )
    ws.conditional_formatting.add(
        f"L{data_start_row}:L{last_row}",
        FormulaRule(formula=[f'L{data_start_row}="完了"'], fill=COMPLETE_FILL, font=COMPLETE_FONT)
    )

    # 遅延 = 赤
    ws.conditional_formatting.add(
        f"H{data_start_row}:H{last_row}",
        FormulaRule(formula=[f'H{data_start_row}="遅延"'], fill=DANGER_FILL, font=DANGER_FONT)
    )
    ws.conditional_formatting.add(
        f"K{data_start_row}:K{last_row}",
        FormulaRule(formula=[f'K{data_start_row}="遅延"'], fill=DANGER_FILL, font=DANGER_FONT)
    )
    ws.conditional_formatting.add(
        f"L{data_start_row}:L{last_row}",
        FormulaRule(formula=[f'L{data_start_row}="遅延"'], fill=DANGER_FILL, font=DANGER_FONT)
    )

    # 予定 = グレー
    ws.conditional_formatting.add(
        f"H{data_start_row}:H{last_row}",
        FormulaRule(formula=[f'H{data_start_row}="予定"'], fill=NEUTRAL_FILL, font=NEUTRAL_FONT)
    )
    ws.conditional_formatting.add(
        f"K{data_start_row}:K{last_row}",
        FormulaRule(formula=[f'K{data_start_row}="予定"'], fill=NEUTRAL_FILL, font=NEUTRAL_FONT)
    )

    # 進行中 = 黄
    ws.conditional_formatting.add(
        f"L{data_start_row}:L{last_row}",
        FormulaRule(formula=[f'L{data_start_row}="進行中"'], fill=WARNING_FILL, font=WARNING_FONT)
    )

    # 列幅設定
    detail_widths = [6, 60, 18, 12, 16, 12, 12, 10, 12, 12, 10, 13]  # L列（基準日参照）を13に拡大
    for i, w in enumerate(detail_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{data_start_row}"

    return data_start_row


def _write_summary_sheet(ws, records, detail_start_row, total_record_count, holidays, team_name="ALL"):
    """進捗サマリーシートを作成（新レイアウト：行3カテゴリヘッダー、行5合計行）

    Returns:
        dict: ダッシュボード参照用の情報
            - data_start_row: データ開始行（6）
            - data_end_row: データ最終行
            - total_row: 合計行（5）
            - ref_row: 基準日行（今日の日付の行）
    """

    # 日付範囲を取得（対象レコードから）
    all_dates = []
    for rec in records:
        for key in ["実施者_予定", "実施者_実績", "検証者_予定", "検証者_実績"]:
            if rec[key]:
                date_obj = _to_date_obj(rec[key])
                if date_obj:
                    all_dates.append(date_obj)

    if not all_dates:
        ws['A1'] = "データがありません"
        return {"data_start_row": 6, "data_end_row": 6, "total_row": 5, "ref_row": 6}

    min_date = min(all_dates)
    max_date = max(all_dates)
    date_range = generate_date_range(min_date, max_date)

    # グリッド線を非表示
    ws.sheet_view.showGridLines = False

    # === 行1: タイトル（ダークネイビー背景） ===
    ws.merge_cells('A1:S1')
    title_cell = ws['A1']
    title_text = f"テスト進捗サマリー（{team_name}）" if team_name != "ALL" else "テスト進捗サマリー（全体）"
    title_cell.value = title_text
    title_cell.font = Font(name="游ゴシック", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=SUMMARY_TITLE_BG, end_color=SUMMARY_TITLE_BG, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # === 行2: 集計期間 + 基準日 ===
    subtitle_fill = PatternFill(start_color=SUMMARY_SUBTITLE_BG, end_color=SUMMARY_SUBTITLE_BG, fill_type="solid")
    # A2:Q2 集計期間
    for col in range(1, 18):  # A-Q
        cell = ws.cell(row=2, column=col)
        cell.fill = subtitle_fill
        cell.border = Border(bottom=THIN_SOLID_SIDE)
    ws['A2'] = f"集計期間: {min_date.strftime('%Y/%m/%d')} ～ {max_date.strftime('%Y/%m/%d')} ({len(date_range)}日間)"
    ws['A2'].font = Font(name="游ゴシック", size=10, color="333333")

    # R2: 基準日ラベル
    ws['R2'] = "基準日:"
    ws['R2'].font = Font(name="游ゴシック", size=10, bold=True, color="333333")
    ws['R2'].fill = subtitle_fill
    ws['R2'].alignment = DATA_ALIGN_RIGHT
    ws['R2'].border = Border(bottom=THIN_SOLID_SIDE)

    # S2: 基準日値（ダッシュボードのB2を参照）
    ws['S2'] = "=ダッシュボード!$B$2"
    ws['S2'].font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
    ws['S2'].fill = PatternFill(start_color="505050", end_color="505050", fill_type="solid")
    ws['S2'].alignment = DATA_ALIGN_CENTER
    ws['S2'].border = Border(bottom=THIN_SOLID_SIDE)
    ws['S2'].number_format = "YYYY/MM/DD"

    ws.row_dimensions[2].height = 22

    # === 行3: カテゴリグループヘッダー（新規追加） ===
    group_font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
    group_align = Alignment(horizontal="center", vertical="center")

    # 共通 (A3:C3)
    ws.merge_cells('A3:C3')
    ws['A3'] = "共通"
    ws['A3'].font = group_font
    ws['A3'].fill = PatternFill(start_color=SUMMARY_GROUP_COMMON, end_color=SUMMARY_GROUP_COMMON, fill_type="solid")
    ws['A3'].alignment = group_align
    ws['A3'].border = Border(top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE, left=THIN_SOLID_SIDE, right=Side(style='medium', color='FFFFFF'))

    # 実施 (D3:J3)
    ws.merge_cells('D3:J3')
    ws['D3'] = "実施"
    ws['D3'].font = group_font
    ws['D3'].fill = PatternFill(start_color=SUMMARY_GROUP_IMPL, end_color=SUMMARY_GROUP_IMPL, fill_type="solid")
    ws['D3'].alignment = group_align
    ws['D3'].border = Border(top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE, left=THIN_SOLID_SIDE, right=Side(style='medium', color='FFFFFF'))

    # 検証 (K3:Q3)
    ws.merge_cells('K3:Q3')
    ws['K3'] = "検証"
    ws['K3'].font = group_font
    ws['K3'].fill = PatternFill(start_color=SUMMARY_GROUP_VERIFY, end_color=SUMMARY_GROUP_VERIFY, fill_type="solid")
    ws['K3'].alignment = group_align
    ws['K3'].border = Border(top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE, left=THIN_SOLID_SIDE, right=Side(style='medium', color='FFFFFF'))

    # 合計 (R3:S3)
    ws.merge_cells('R3:S3')
    ws['R3'] = "合計"
    ws['R3'].font = group_font
    ws['R3'].fill = PatternFill(start_color=SUMMARY_GROUP_TOTAL, end_color=SUMMARY_GROUP_TOTAL, fill_type="solid")
    ws['R3'].alignment = group_align
    ws['R3'].border = Border(top=THIN_SOLID_SIDE, bottom=THIN_SOLID_SIDE, left=THIN_SOLID_SIDE, right=THIN_SOLID_SIDE)

    ws.row_dimensions[3].height = 24

    # === 行4: サブヘッダー（カテゴリ別配色） ===
    header_row = 4
    summary_headers = [
        "日付", "曜", "営業日",
        "予定", "実績", "消化率", "予定累計", "実績累計", "累計消化率", "判定",
        "予定", "実績", "消化率", "予定累計", "実績累計", "累計消化率", "判定",
        "予定", "実績",
    ]

    # カテゴリ別の背景色マッピング
    header_fills = {
        (1, 3): SUMMARY_SUB_COMMON,   # A-C: 共通
        (4, 10): SUMMARY_SUB_IMPL,    # D-J: 実施
        (11, 17): SUMMARY_SUB_VERIFY, # K-Q: 検証
        (18, 19): SUMMARY_SUB_TOTAL,  # R-S: 合計
    }

    def get_header_fill(col):
        for (start, end), color in header_fills.items():
            if start <= col <= end:
                return PatternFill(start_color=color, end_color=color, fill_type="solid")
        return HEADER_FILL

    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
        cell.fill = get_header_fill(col)
        cell.alignment = HEADER_ALIGN
        # 下辺は二重線
        cell.border = Border(
            top=THIN_SOLID_SIDE,
            bottom=DOUBLE_SIDE,
            left=MEDIUM_SOLID_SIDE if col in (1, 4, 11, 18) else THIN_SOLID_SIDE,
            right=MEDIUM_SOLID_SIDE if col in (3, 10, 17, 19) else THIN_SOLID_SIDE,
        )

    ws.row_dimensions[4].height = 22

    # === 行5: 合計サマリー行 ===
    summary_row = 5
    data_start_row = 6
    detail_last_row = detail_start_row + total_record_count - 1
    last_data_row = data_start_row + len(date_range) - 1

    summary_fill = PatternFill(start_color=SUMMARY_SUMMARY_ROW_BG, end_color=SUMMARY_SUMMARY_ROW_BG, fill_type="solid")
    summary_font = Font(name="游ゴシック", size=11, bold=True)

    # A5: "合計"
    ws.cell(row=summary_row, column=1, value="合計").font = summary_font
    ws.cell(row=summary_row, column=1).fill = summary_fill
    ws.cell(row=summary_row, column=1).alignment = DATA_ALIGN_CENTER

    # B5, C5: 空
    for col in [2, 3]:
        cell = ws.cell(row=summary_row, column=col)
        cell.fill = summary_fill
        cell.font = summary_font

    # チームフィルタ用のCOUNTIFS条件
    if team_name == "ALL":
        team_condition = ""
    else:
        team_condition = f',明細!$D${detail_start_row}:$D${detail_last_row},"{team_name}"'

    # 合計行の数式（行6〜最終行のSUM）
    # D5: 実施_予定合計
    ws.cell(row=summary_row, column=4, value=f"=SUM(D{data_start_row}:D{last_data_row})").font = summary_font
    # E5: 実施_実績合計
    ws.cell(row=summary_row, column=5, value=f"=SUM(E{data_start_row}:E{last_data_row})").font = summary_font
    # F5: 実施_消化率
    ws.cell(row=summary_row, column=6, value=f"=IF(D5=0,0,E5/D5)").font = summary_font
    ws.cell(row=summary_row, column=6).number_format = "0.0%"
    # G5: 実施_予定累計（合計と同値）
    ws.cell(row=summary_row, column=7, value=f"=D5").font = summary_font
    # H5: 実施_実績累計（合計と同値）
    ws.cell(row=summary_row, column=8, value=f"=E5").font = summary_font
    # I5: 実施_累計消化率
    ws.cell(row=summary_row, column=9, value=f'=IF(G5=0,"",H5/G5)').font = summary_font
    ws.cell(row=summary_row, column=9).number_format = "0.0%"
    # J5: 実施_判定
    ws.cell(row=summary_row, column=10, value=f'=IF(I5>=1,"完了",IF(H5>=G5,"順調","遅延"))').font = summary_font

    # K5: 検証_予定合計
    ws.cell(row=summary_row, column=11, value=f"=SUM(K{data_start_row}:K{last_data_row})").font = summary_font
    # L5: 検証_実績合計
    ws.cell(row=summary_row, column=12, value=f"=SUM(L{data_start_row}:L{last_data_row})").font = summary_font
    # M5: 検証_消化率
    ws.cell(row=summary_row, column=13, value=f"=IF(K5=0,0,L5/K5)").font = summary_font
    ws.cell(row=summary_row, column=13).number_format = "0.0%"
    # N5: 検証_予定累計
    ws.cell(row=summary_row, column=14, value=f"=K5").font = summary_font
    # O5: 検証_実績累計
    ws.cell(row=summary_row, column=15, value=f"=L5").font = summary_font
    # P5: 検証_累計消化率
    ws.cell(row=summary_row, column=16, value=f'=IF(N5=0,"",O5/N5)').font = summary_font
    ws.cell(row=summary_row, column=16).number_format = "0.0%"
    # Q5: 検証_判定
    ws.cell(row=summary_row, column=17, value=f'=IF(P5>=1,"完了",IF(O5>=N5,"順調","遅延"))').font = summary_font

    # R5: 合計_予定
    ws.cell(row=summary_row, column=18, value=f"=D5+K5").font = summary_font
    # S5: 合計_実績
    ws.cell(row=summary_row, column=19, value=f"=E5+L5").font = summary_font

    # 合計行のスタイル適用
    for col in range(1, 20):
        cell = ws.cell(row=summary_row, column=col)
        cell.fill = summary_fill
        cell.alignment = DATA_ALIGN_CENTER
        # 上辺: 二重線、下辺: medium
        cell.border = Border(
            top=DOUBLE_SIDE,
            bottom=MEDIUM_SOLID_SIDE,
            left=MEDIUM_SOLID_SIDE if col in (1, 4, 11, 18) else THIN_SOLID_SIDE,
            right=MEDIUM_SOLID_SIDE if col in (3, 10, 17, 19) else THIN_SOLID_SIDE,
        )
        # カウント列は#,##0書式
        if col in (4, 5, 7, 8, 11, 12, 14, 15, 18, 19):
            cell.number_format = "#,##0"

    ws.row_dimensions[5].height = 24

    # === 行6以降: データ行 ===
    for i, date_obj in enumerate(date_range):
        row = data_start_row + i

        # 曜日・営業日は数式で計算
        weekday_formula = f'=CHOOSE(WEEKDAY(A{row},2),"月","火","水","木","金","土","日")'
        business_formula = f'=IF(OR(WEEKDAY(A{row},2)>=6,COUNTIF(祝日マスタ!$A:$A,A{row})>0),"非営業日","営業日")'

        # 明細シートを参照するCOUNTIFS関数
        jisshi_yotei = f'=COUNTIFS(明細!$F${detail_start_row}:$F${detail_last_row},A{row}{team_condition})'
        jisshi_jisseki = f'=COUNTIFS(明細!$G${detail_start_row}:$G${detail_last_row},A{row}{team_condition})'
        jisshi_rate = f'=IF(D{row}=0,0,E{row}/D{row})'  # 0件時は0を返す（指示書準拠）
        jisshi_yotei_cum = f'=SUM($D${data_start_row}:D{row})'
        jisshi_jisseki_cum = f'=SUM($E${data_start_row}:E{row})'
        jisshi_cum_rate = f'=IF(G{row}=0,"",H{row}/G{row})'
        jisshi_status = f'=IF(A{row}>$S$2,"予定",IF(G{row}=0,"－",IF(I{row}>=1,"完了",IF(H{row}>=G{row},"順調","遅延"))))'

        kensho_yotei = f'=COUNTIFS(明細!$I${detail_start_row}:$I${detail_last_row},A{row}{team_condition})'
        kensho_jisseki = f'=COUNTIFS(明細!$J${detail_start_row}:$J${detail_last_row},A{row}{team_condition})'
        kensho_rate = f'=IF(K{row}=0,0,L{row}/K{row})'  # 0件時は0を返す（指示書準拠）
        kensho_yotei_cum = f'=SUM($K${data_start_row}:K{row})'
        kensho_jisseki_cum = f'=SUM($L${data_start_row}:L{row})'
        kensho_cum_rate = f'=IF(N{row}=0,"",O{row}/N{row})'
        kensho_status = f'=IF(A{row}>$S$2,"予定",IF(N{row}=0,"－",IF(P{row}>=1,"完了",IF(O{row}>=N{row},"順調","遅延"))))'

        total_yotei = f'=D{row}+K{row}'
        total_jisseki = f'=E{row}+L{row}'

        values = [
            date_obj,           # A: 日付（日付オブジェクト）
            weekday_formula,    # B: 曜日
            business_formula,   # C: 営業日
            jisshi_yotei,       # D: 実施_予定
            jisshi_jisseki,     # E: 実施_実績
            jisshi_rate,        # F: 実施_消化率
            jisshi_yotei_cum,   # G: 実施_予定累計
            jisshi_jisseki_cum, # H: 実施_実績累計
            jisshi_cum_rate,    # I: 実施_累計消化率
            jisshi_status,      # J: 実施_判定
            kensho_yotei,       # K: 検証_予定
            kensho_jisseki,     # L: 検証_実績
            kensho_rate,        # M: 検証_消化率
            kensho_yotei_cum,   # N: 検証_予定累計
            kensho_jisseki_cum, # O: 検証_実績累計
            kensho_cum_rate,    # P: 検証_累計消化率
            kensho_status,      # Q: 検証_判定
            total_yotei,        # R: 合計_予定
            total_jisseki,      # S: 合計_実績
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = DATA_FONT

            # カテゴリ境界にmedium罫線、それ以外はthin/dotted
            left_medium = col in (4, 11, 18)  # D, K, R
            right_medium = col in (3, 10, 17, 19)  # C, J, Q, S
            left_solid = col in (1, 4, 7, 10, 11, 14, 17, 18)
            right_solid = col in (3, 6, 9, 10, 13, 16, 17, 19)

            cell.border = Border(
                left=MEDIUM_SOLID_SIDE if left_medium else (THIN_SOLID_SIDE if left_solid else THIN_DOTTED_SIDE),
                right=MEDIUM_SOLID_SIDE if right_medium else (THIN_SOLID_SIDE if right_solid else THIN_DOTTED_SIDE),
                top=THIN_SOLID_SIDE,
                bottom=THIN_SOLID_SIDE,
            )

            cell.alignment = DATA_ALIGN_CENTER

            # 数値書式
            if col == 1:  # 日付列
                cell.number_format = "YYYY/MM/DD"
            elif col in (4, 5, 7, 8, 11, 12, 14, 15, 18, 19):  # カウント列
                cell.number_format = "#,##0"
            elif col in (6, 9, 13, 16):  # 消化率列
                cell.number_format = "0.0%"

    # === 条件付き書式 ===
    # ステータス条件付き書式（J5:J{last}, Q5:Q{last}）- 新配色
    for status_col in ['J', 'Q']:
        for status, colors in STATUS_COLORS.items():
            ws.conditional_formatting.add(
                f"{status_col}{summary_row}:{status_col}{last_data_row}",
                CellIsRule(
                    operator='equal',
                    formula=[f'"{status}"'],
                    fill=PatternFill(start_color=colors["bg"], end_color=colors["bg"], fill_type="solid"),
                    font=Font(color=colors["fg"], bold=colors["bold"])
                )
            )

    # データバー（I列、P列） ※行6から（行5の合計行を除く）
    # グラデーションを逆にするため、Ruleを直接作成
    from openpyxl.formatting.rule import Rule, DataBar, FormatObject
    from openpyxl.styles import Color

    # 実施用データバー（青系）- グラデーション無効化（単色）
    impl_databar = DataBar(
        minLength=0,
        maxLength=100,
        showValue=True,
        cfvo=[FormatObject(type='num', val=0), FormatObject(type='num', val=1)],
        color=Color(rgb="FF" + DATABAR_IMPL)
    )
    impl_rule = Rule(type='dataBar', dataBar=impl_databar)
    ws.conditional_formatting.add(f"I{data_start_row}:I{last_data_row}", impl_rule)

    # 検証用データバー（緑系）- グラデーション無効化（単色）
    verify_databar = DataBar(
        minLength=0,
        maxLength=100,
        showValue=True,
        cfvo=[FormatObject(type='num', val=0), FormatObject(type='num', val=1)],
        color=Color(rgb="FF" + DATABAR_VERIFY)
    )
    verify_rule = Rule(type='dataBar', dataBar=verify_databar)
    ws.conditional_formatting.add(f"P{data_start_row}:P{last_data_row}", verify_rule)

    # 基準日行ハイライト（A6:S{last}）
    ws.conditional_formatting.add(
        f"A{data_start_row}:S{last_data_row}",
        FormulaRule(
            formula=[f'$A{data_start_row}=$S$2'],
            fill=PatternFill(start_color=BASEDATE_HIGHLIGHT_BG, end_color=BASEDATE_HIGHLIGHT_BG, fill_type="solid"),
            font=Font(color=BASEDATE_HIGHLIGHT_FG, bold=True)
        )
    )

    # 非営業日は行全体を薄いグレーに（C列="非営業日"）
    ws.conditional_formatting.add(
        f"A{data_start_row}:S{last_data_row}",
        FormulaRule(
            formula=[f'$C{data_start_row}="非営業日"'],
            fill=NEUTRAL_FILL
        )
    )

    # === 列幅設定 ===
    for col_letter, width in SUMMARY_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # === フリーズ（行5まで、C列まで固定） ===
    ws.freeze_panes = 'D6'

    # 基準日行を特定
    today = datetime.now().date()
    ref_row = data_start_row  # デフォルト
    for i, date_obj in enumerate(date_range):
        if date_obj.date() == today:
            ref_row = data_start_row + i
            break

    return {
        "data_start_row": data_start_row,
        "data_end_row": last_data_row,
        "total_row": summary_row,
        "ref_row": ref_row,
    }


# ===================================================================
#  メイン処理
# ===================================================================

def main():
    print("=" * 60)
    print("  テスト予定・実績 集計スクリプト v4")
    print("=" * 60)

    # --- CLI引数パース ---
    parser = argparse.ArgumentParser(description="テスト予定・実績 集計スクリプト")
    parser.add_argument("folder", nargs="?", help="対象フォルダパス")
    parser.add_argument("-o", "--output", help="出力ファイルパス")
    parser.add_argument("-s", "--subfolders", action="store_true", default=True,
                        help="サブフォルダを含める（デフォルト: True）")
    parser.add_argument("--no-subfolders", action="store_true",
                        help="サブフォルダを含めない")
    args = parser.parse_args()

    # CLIモード or GUIウィザードモード
    cli_mode = False
    week_from = None
    week_to = None

    if args.folder:
        # CLIモード
        cli_mode = True
        folder_path = args.folder
        if args.output:
            output_path = args.output
        else:
            # デフォルト出力パス
            output_path = os.path.join("output", f"test_progress_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        include_subfolders = not args.no_subfolders
        # CLIモードでは週範囲は指定しない（将来的に引数追加可能）
    else:
        # --- ウィザードUI実行 ---
        config = run_wizard()

        if config is None:
            print("\n  キャンセルされました。")
            sys.exit(0)

        folder_path = config["folder_path"]
        output_path = config["output_path"]
        include_subfolders = config["include_subfolders"]
        week_from = config.get("week_from")
        week_to = config.get("week_to")

    # キャッシュファイルのパス（出力ファイルと同じディレクトリ）
    cache_dir = os.path.dirname(output_path)
    cache_file = os.path.join(cache_dir, ".test_collector_cache.json")

    print(f"\n  対象フォルダ: {folder_path}")
    print(f"  サブフォルダ: {'含める' if include_subfolders else '含めない'}")
    print(f"  出力先:       {output_path}")
    print(f"  対象シート:   {SHEET_PREFIX}* で始まるシート\n")

    records = collect_data(folder_path, cache_file, include_subfolders)

    if not records:
        msg = (
            "データが見つかりませんでした。\n"
            "・フォルダパスを確認してください\n"
            f"・シート名が「{SHEET_PREFIX}」で始まるか確認してください\n"
            "・C列にテストIDが入っているか確認してください"
        )
        print(f"\n  ⚠ {msg}")
        if not cli_mode:
            root_err = tk.Tk()
            root_err.withdraw()
            root_err.attributes("-topmost", True)
            messagebox.showwarning("データなし", msg)
            root_err.destroy()
        sys.exit(1)

    write_excel(records, output_path, week_from=week_from, week_to=week_to)

    print("\n" + "=" * 60)

    # 完了メッセージ
    team_counts = defaultdict(int)
    for rec in records:
        team_counts[rec["チーム名"]] += 1

    team_info = "\n".join([f"  - {team}: {count}件" for team, count in sorted(team_counts.items())])

    print(f"\n  集計完了！")
    print(f"  出力先: {output_path}")
    print(f"  明細: {len(records)}件")
    print(f"  チーム別内訳:")
    for team, count in sorted(team_counts.items()):
        print(f"    - {team}: {count}件")

    if not cli_mode:
        root_done = tk.Tk()
        root_done.withdraw()
        root_done.attributes("-topmost", True)
        messagebox.showinfo(
            "完了",
            f"集計が完了しました！\n\n"
            f"出力先:\n{output_path}\n\n"
            f"明細: {len(records)}件\n\n"
            f"チーム別内訳:\n{team_info}",
        )
        root_done.destroy()


if __name__ == "__main__":
    main()
